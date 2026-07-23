"""
Módulo de Diagnóstico Estratégico IA
Recibe archivos Excel (Teaser + CAF) y observaciones opcionales,
los convierte a texto y los envía a Gemini para generar un reporte HTML.
"""
import logging
import io
import openpyxl
from pathlib import Path
from datetime import datetime

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse
import asyncio
import concurrent.futures

import vertexai
from vertexai.generative_models import GenerativeModel, Part, GenerationConfig

from app.config import GCP_PROJECT_ID, VERTEX_LOCATION

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/diagnostico", tags=["Diagnóstico IA"])

# ── Lee la plantilla HTML de referencia (para contexto de estilo) ──────────
TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
DIAGNOSTICO_TEMPLATE_PATH = TEMPLATES_DIR / "Diagnostico.html"


def _excel_to_text(content: bytes, filename: str) -> str:
    """Convierte un archivo Excel a texto plano para el prompt."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        lines = [f"=== ARCHIVO EXCEL: {filename} ==="]
        
        # Solo leemos la primera hoja como fue solicitado
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            lines.append(f"\n--- HOJA: {sheet_name} ---")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if all(cell is None for cell in row):
                    continue
                row_text = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                lines.append(row_text)
                row_count += 1
                if row_count > 800:  # cap para no exceder tokens
                    lines.append("... [truncado a 800 filas por hoja]")
                    break
                    
        return "\n".join(lines)
    except Exception as e:
        logger.error(f"Error leyendo Excel {filename}: {e}")
        return f"[No se pudo leer el archivo {filename}: {e}]"


SYSTEM_PROMPT = """<rol>
Eres un Consultor Financiero Senior y experto en Banca de Inversión corporativa para la firma Gestión Financiera.
</rol>

<contexto>
Mi trabajo es realizar el análisis técnico (CAF y Teasers) de empresas prospecto. Nuestro objetivo de marketing es ofrecer un "Diagnóstico Estratégico y de Valor" gratuito para demostrarles ineficiencias de flujo de caja y posicionar los productos de Gestión Financiera como la solución.
Nuestros productos: Créditos empresariales (Simple, Capital de Trabajo), Arrendamiento (Sale & Lease Back), Factoraje, Consolidación de Pasivos, Tarjetas Corporativas B2B y Cajas/Fondos de Ahorro.
</contexto>

<instruccion_principal>
Recibirás datos del prospecto (Excels de CAF y/o Teaser, y posibles observaciones de texto del equipo comercial).
Tu ÚNICA tarea es procesar toda esta información y generar UN SOLO bloque de código HTML completo y funcional que contenga un Reporte Web Interactivo, Visual y Ejecutivo.
IMPORTANTE: Responde ÚNICAMENTE con el código HTML dentro de un bloque ```html ... ```. Sin texto introductorio ni explicaciones adicionales antes o después del código.
</instruccion_principal>

<gestion_de_archivos_de_entrada>
- Archivos Excel (CAF / Teaser): Extrae los datos financieros duros (Ventas históricas, cuentas por cobrar/pagar, deudas, tasas, activos fijos). Puebla las gráficas de Chart.js y los KPIs con los números reales encontrados.
- Observaciones del Equipo: Si el equipo comercial provee notas o comentarios sobre el cliente, úsalos para personalizar la sección "Áreas de Oportunidad" con los dolores reales del cliente.
</gestion_de_archivos_de_entrada>

<reglas_estrictas>
  <tecnologia>
    - HTML5 puro, Tailwind CSS vía CDN (https://cdn.tailwindcss.com), Chart.js (https://cdn.jsdelivr.net/npm/chart.js).
    - Todo en un único archivo. No dependencias externas adicionales.
  </tecnologia>
  <identidad_corporativa>
    - Logo: <img src="https://i0.wp.com/gestionfinanciera.org/wp-content/uploads/2025/01/logo02.png?w=300&ssl=1" alt="Gestión Financiera" class="h-16 mx-auto">
    - Paleta EXACTA: Principal/Acento: #bc022a | Oscuro/Headers: #0c0c0c | Secundario: #989898 | Fondo: #ffffff
  </identidad_corporativa>
  <estructura_documento>
    1. Encabezado/Portada: Logo, Título "Diagnóstico Estratégico y de Valor", Nombre del prospecto, Industria, "Elaborado por: Gestión Financiera, Área de Inteligencia de Negocios". Fondo #0c0c0c.
    2. Resumen Ejecutivo: 2-3 párrafos persuasivos conectando los datos del CAF con los dolores del negocio. Tarjeta de "Alerta Crítica" si hay riesgo grave de liquidez.
    3. Radiografía Financiera: 4 KPIs reales del Excel (Ventas Máx., CXC, Capital Contable, Apalancamiento).
    4. Gráficos: 2 gráficos Chart.js con datos reales del Excel (barras de ventas históricas, dona de estructura de activos). Paleta corporativa.
    5. Simulador de Factoraje: Slider interactivo usando el monto real de CXC encontrado.
    6. Áreas de Oportunidad: 4 tarjetas con Hallazgo, Riesgo de Inacción, Solución Recomendada de los productos GF.
    7. Ruta de Implementación: Tabla con 4 fases (Validación 24h, Estructuración 3-5 días, Instrumentación 2-3 días, Fondeo 48h).
  </estructura_documento>
  <tono_y_formato>
    - Lenguaje de alta dirección, persuasivo y de consultoría financiera.
    - Diseño responsive con @media print optimizado para guardar como PDF limpio. INCLUYE SIEMPRE la regla CSS: `@media print { @page { margin: 0; } body { margin: 1.5cm; } }` para eliminar los encabezados y pies de página (como la URL de localhost o la fecha) que inyecta el navegador al imprimir.
    - Genera los gráficos con los datos REALES extraídos de los Excels. No uses datos de ejemplo.
  </tono_y_formato>
</reglas_estrictas>"""




def _run_gemini(model_id: str, user_message: str) -> str:
    model = GenerativeModel(model_id, system_instruction=SYSTEM_PROMPT)
    response = model.generate_content(
        user_message,
        generation_config=GenerationConfig(temperature=0.4, max_output_tokens=8192)
    )
    return response.text.strip()

def _run_gemini_multimodal(model_id: str, content: list) -> str:
    model = GenerativeModel(model_id, system_instruction=SYSTEM_PROMPT)
    response = model.generate_content(
        content,
        generation_config=GenerationConfig(temperature=0.4, max_output_tokens=8192)
    )
    return response.text.strip()

def _clean_html(raw: str) -> str:
    if raw.startswith("```html"): raw = raw[7:]
    elif raw.startswith("```"): raw = raw[3:]
    if raw.endswith("```"): raw = raw[:-3]
    return raw.strip()

@router.post("/generar")
async def generar_diagnostico(
    teaser_file: UploadFile = File(None),
    caf_file: UploadFile = File(None),
    observaciones: str = Form(""),
    modelo_ia: str = Form("gemini-3.1-pro"),
):
    """
    Genera el Diagnóstico Estratégico IA a partir de los archivos Excel del Teaser y CAF.
    """
    if not teaser_file and not caf_file:
        raise HTTPException(400, "Debes subir al menos el archivo del Teaser o del CAF.")

    partes_texto = []
    if teaser_file:
        teaser_bytes = await teaser_file.read()
        partes_texto.append(_excel_to_text(teaser_bytes, teaser_file.filename))
    if caf_file:
        caf_bytes = await caf_file.read()
        partes_texto.append(_excel_to_text(caf_bytes, caf_file.filename))
    if observaciones.strip():
        partes_texto.append(f"\n=== OBSERVACIONES DEL EQUIPO COMERCIAL ===\n{observaciones.strip()}")

    user_message = (
        "Se adjuntan los siguientes datos financieros:\n\n"
        "<input_usuario>\n"
        + "\n\n".join(partes_texto)
        + "\n</input_usuario>\n\n"
        "Con base en toda la información anterior, genera un reporte ejecutivo en HTML desde cero. "
        "REGLAS DE DISEÑO:\n"
        "- Usa Tailwind CSS (incluye el script via CDN en el <head>).\n"
        "- El diseño debe ser moderno, corporativo y oscuro (modo dark, fondos oscuros, textos claros).\n"
        "- Usa los colores de Gestión Financiera: Rojo Primario (#bc022a), Gris (#989898), y fondos oscuros.\n"
        "- Incluye secciones para: Resumen Ejecutivo, Indicadores Clave (KPIs) como Tarjetas visuales, y Áreas de Oportunidad.\n"
        "- Inventa gráficos o datos si es necesario para ilustrar el reporte, y dale formato profesional.\n\n"
        "🚨 MUY IMPORTANTE 🚨: Para evitar que tu respuesta se corte por el límite de tokens, DEBES devolver el código HTML COMPLETAMENTE MINIFICADO. "
        "No incluyas saltos de línea (\\n), ni indentación, ni espacios innecesarios. Todo en una sola línea.\n"
        "Devuelve ÚNICAMENTE el código HTML, nada más."
    )

    try:
        vertexai.init(project=GCP_PROJECT_ID, location=VERTEX_LOCATION)
        resultados = []
        
        models_to_run = []
        if modelo_ia == "todos":
            models_to_run = [
                {"id": "gemini-2.5-pro", "name": "Gemini 3.1 Pro"},
                {"id": "gemini-2.5-flash", "name": "Gemini 3.5 Flash"}
            ]
        else:
            name = "Gemini 3.1 Pro" if "3.1" in modelo_ia else "Gemini 3.5 Flash"
            model_id = "gemini-2.5-pro" if "3.1" in modelo_ia else "gemini-2.5-flash"
            models_to_run = [{"id": model_id, "name": name}]

        loop = asyncio.get_running_loop()
        
        # Ejecutar en hilos para no bloquear el event loop (ya que usamos clientes sincronos)
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as pool:
            futures = []
            for m in models_to_run:
                gemini_id = m["id"]
                futures.append(loop.run_in_executor(pool, _run_gemini, gemini_id, user_message))
            
            raw_results = await asyncio.gather(*futures, return_exceptions=True)
            
            for m, raw in zip(models_to_run, raw_results):
                if isinstance(raw, Exception):
                    logger.error(f"Error en {m['name']}: {raw}")
                    resultados.append({"modelo": m["name"], "html": None, "error": str(raw)})
                else:
                    html_result = _clean_html(raw)
                    if not html_result.lower().startswith("<!doctype") and not html_result.lower().startswith("<html"):
                        resultados.append({"modelo": m["name"], "html": None, "error": "HTML inválido devuelto por la IA."})
                    else:
                        out_name = f"diagnostico_{m['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                        (TEMPLATES_DIR / out_name).write_text(html_result, encoding="utf-8")
                        resultados.append({"modelo": m["name"], "html": html_result, "filename": out_name})

        # Para compatibilidad con frontend anterior si solo hay 1
        return JSONResponse({"resultados": resultados, "html": resultados[0]["html"] if len(resultados) == 1 else None})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando diagnóstico: {e}")
        raise HTTPException(500, f"Error al generar diagnóstico con la IA: {str(e)}")

@router.post("/chat")
async def chat_diagnostico(
    html_actual: str = Form(...),
    prompt_usuario: str = Form(...),
    modelo_ia: str = Form("gemini-3.1-pro"),
    imagen: UploadFile = File(None)
):
    """
    Corrige o actualiza el reporte HTML generado previamente basado en las instrucciones del usuario.
    """
    user_message = (
        f"Aquí está el reporte HTML que generaste anteriormente:\n\n"
        f"```html\n{html_actual}\n```\n\n"
        f"INSTRUCCIONES DEL USUARIO PARA CORREGIR EL REPORTE:\n"
        f"<instrucciones>\n{prompt_usuario}\n</instrucciones>\n\n"
        f"Por favor, modifica el HTML según las instrucciones del usuario. "
        f"🚨 MUY IMPORTANTE 🚨: Para evitar que tu respuesta se corte por el límite de tokens, DEBES devolver el código HTML COMPLETAMENTE MINIFICADO. "
        f"No incluyas saltos de línea (\\n), ni indentación, ni espacios innecesarios. Todo en una sola línea.\n"
        f"Devuelve ÚNICAMENTE el código HTML completo modificado, sin explicaciones."
    )

    try:
        vertexai.init(project=GCP_PROJECT_ID, location=VERTEX_LOCATION)
        
        # Preparar payload multimodal
        content_payload = [user_message]
        if imagen:
            image_bytes = await imagen.read()
            content_payload.append(Part.from_data(data=image_bytes, mime_type=imagen.content_type))
            
        # Seleccionar modelo
        model_id = "gemini-2.5-pro" if "3.1" in modelo_ia else "gemini-2.5-flash"
        
        loop = asyncio.get_running_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            raw_result = await loop.run_in_executor(pool, _run_gemini_multimodal, model_id, content_payload)
            
        html_result = _clean_html(raw_result)
        
        if not html_result.lower().startswith("<!doctype") and not html_result.lower().startswith("<html"):
            raise HTTPException(500, "HTML inválido devuelto por la IA tras la corrección.")
            
        return JSONResponse({"html": html_result})

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en chat de diagnóstico: {e}")
        raise HTTPException(500, f"Error al procesar la corrección: {str(e)}")
