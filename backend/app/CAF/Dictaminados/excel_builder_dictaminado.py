import logging
import re
import uuid
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
import io
import base64
import unicodedata
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FONT   = Font(bold=True, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="4CAF50")
INPUT_FILL    = PatternFill("solid", fgColor="FFF9C4")
NOTA_FILL     = PatternFill("solid", fgColor="1565C0")   # azul oscuro para headers de nota
NOTA_FONT     = Font(bold=True, color="FFFFFF", size=10)
NOTA_ROW_FILL = PatternFill("solid", fgColor="DDEEFF")   # azul celeste suave para filas de nota
NOTA_END_FILL = PatternFill("solid", fgColor="B3D4F5")   # azul un poco más intenso para cierre
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
_OCR_NOISE = {".", ",", "_", "|", "/", "\\", ":", ";", "I", "l"}

# Secciones del Balance según fila en la plantilla
BALANCE_SECTIONS = [
    (range(9, 19),  "ACTIVO CIRCULANTE",  "1A7A4A"),
    (range(24, 32), "ACTIVO FIJO",         "2E7D32"),
    (range(37, 40), "ACTIVO DIFERIDO",     "388E3C"),
    (range(52, 58), "PASIVO CIRCULANTE",   "C62828"),
    (range(64, 66), "PASIVO LARGO PLAZO",  "E53935"),
    (range(74, 79), "CAPITAL CONTABLE",    "6A1B9A"),
]
EDO_SECTIONS = [
    (range(7, 9),   "INGRESOS",            "1A7A4A"),
    (range(8, 12),  "COSTOS Y GASTOS",     "C62828"),
    (range(15, 18), "RESULTADO FINANCIERO","1565C0"),
    (range(20, 22), "OTROS",               "6D4C41"),
    (range(24, 25), "IMPUESTOS",           "4E342E"),
    (range(29, 30), "DEPRECIACIÓN",        "37474F"),
]

def _get_section(row_num: int, sheet_name: str) -> tuple[str, str]:
    """Retorna (Label, ColorHex) según la fila del template donde caiga."""
    sections = BALANCE_SECTIONS if sheet_name == "Balance" else EDO_SECTIONS
    for (rng, label, color) in sections:
        if row_num in rng:
            return label, color
    return None, None


def _clean_text(text: str) -> str:
    """Elimina acentos, puntuación extra y convierte a minúsculas para comparaciones."""
    if not text:
        return ""
    text = str(text).lower()
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^a-z0-9\s]', '', text)
    return text.strip()


def _parse_monto(monto_str: str) -> float:
    """Convierte un string de monto a float (manejando $, comas y paréntesis para negativos)."""
    if not monto_str:
        return 0.0
    val_str = str(monto_str).strip()
    is_negative = '(' in val_str and ')' in val_str
    
    # Remove all non-numeric chars except digits and decimal point
    clean_str = re.sub(r'[^\d\.]', '', val_str)
    if not clean_str:
        return 0.0
    
    try:
        val = float(clean_str)
        return -val if is_negative else val
    except ValueError:
        return 0.0


def _find_best_match(target: str, flat_data: list, threshold: float = 0.8) -> str:
    """
    Busca la mejor coincidencia del 'target' en 'flat_data' (lista de tuplas (concepto, monto)).
    Retorna el monto si la coincidencia supera el threshold, o None.
    """
    if not target or not flat_data:
        return None
        
    target_clean = _clean_text(target)
    best_score = 0.0
    best_monto = None
    
    for concept, monto in flat_data:
        concept_clean = _clean_text(concept)
        if not concept_clean:
            continue
            
        score = SequenceMatcher(None, target_clean, concept_clean).ratio()
        if score > best_score:
            best_score = score
            best_monto = monto
            
    if best_score >= threshold:
        return best_monto
    return None

# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_numeric(text: str) -> bool:
    cleaned = re.sub(r'[\$,\s\(\)]', '', str(text))
    try:
        float(cleaned)
        return True
    except ValueError:
        return False

def _clean_monto_ocr(text: str) -> str:
    """
    Limpia alucinaciones de OCR en los montos (ej. '9,231,513 AAAAAA69' -> '9,231,513').
    Si no contiene números (ej. 'detallan.', 'Acreedores'), devuelve string vacío.
    """
    if not text:
        return ""
    # Busca el primer bloque que parezca un número financiero: $, números, comas, puntos, paréntesis.
    # Exige al menos un dígito (\d) para no hacer match con puros puntos o comas.
    m = re.search(r'(?:[\$\(]\s*)?(?:[\d,\.]*\d[\d,\.]*)(?:\s*\))?', text)
    if m:
        return m.group(0).strip()
    
    # Conservar guión si es lo único
    if text.strip() == "-":
        return "-"
        
    return ""



def _tokenize_cells(row) -> list:
    """Extrae tokens de las celdas de una fila, de izquierda a derecha."""
    tokens = []
    lines_per_cell = []
    for c in row:
        if not c or not c.get("text"):
            continue
        cell_lines = [line.strip() for line in str(c["text"]).strip().split('\n')]
        lines_per_cell.append(cell_lines)

    if not lines_per_cell:
        return []

    max_lines = max(len(lines) for lines in lines_per_cell)
    for i in range(max_lines):
        for cell_lines in lines_per_cell:
            if i < len(cell_lines) and cell_lines[i]:
                tokens.append(cell_lines[i])
    return tokens


def _extract_pairs_from_native_cells(row) -> list:
    """
    For rows coming from DocAI native table detection (notas_dictaminado mode).

    CLAVE: Usa POSICIÓN de celda, no detección numérica:
      - cells[0]     = Concepto (primera celda, siempre texto)
      - cells[-2]    = Monto Año 1 (penúltima = año más reciente)
      - cells[-1]    = Monto Año 2 (última = año anterior)

    Esto es más robusto porque DocAI ya separó las celdas correctamente.
    Omite filas de un solo token (totalmente vacías o solo símbolo).
    """
    if not row:
        return []

    # Get only non-empty cells
    cells = [c for c in row if c.get("text", "").strip()]
    if len(cells) < 2:
        return []

    # Encontrar la primera celda que tiene un monto numérico válido
    first_num_idx = -1
    for i, c in enumerate(cells):
        if _clean_monto_ocr(c["text"].strip()):
            first_num_idx = i
            break
            
    if first_num_idx == -1:
        return [] # No hay números en toda la fila, descartar
        
    if first_num_idx == 0:
        concepto = "Dato"
    else:
        concepto = " ".join(c["text"].strip() for c in cells[:first_num_idx])
        
    data_cells = cells[first_num_idx:]
    pairs = []
    
    # Si hay más de 2 datos numéricos, pivoteamos (recortamos) la tabla hacia abajo
    if len(data_cells) > 2 and len(data_cells) % 2 == 0:
        # Probablemente simétrico (Año 1 y Año 2)
        half = len(data_cells) // 2
        for i in range(half):
            m1 = _clean_monto_ocr(data_cells[i]["text"].strip())
            m2 = _clean_monto_ocr(data_cells[i + half]["text"].strip())
            
            if m1 or m2:
                # Omitir encabezados de años
                try:
                    v1 = float(re.sub(r'[\$,\s\(\)]', '', m1)) if m1 and m1 != "-" else 0
                    v2 = float(re.sub(r'[\$,\s\(\)]', '', m2)) if m2 and m2 != "-" else 0
                    if m1 and m2 and 1990 <= v1 <= 2030 and 1990 <= v2 <= 2030:
                        continue
                except (ValueError, AttributeError):
                    pass
                    
                suffix = f" (Dato {i+1})" if half > 1 else ""
                pairs.append((f"{concepto}{suffix}", m1, m2))
    else:
        # 1 o 2 celdas, o impar (no simétrico)
        # Si son 2, es el comportamiento normal
        if len(data_cells) == 2:
            m1 = _clean_monto_ocr(data_cells[0]["text"].strip())
            m2 = _clean_monto_ocr(data_cells[1]["text"].strip())
            try:
                v1 = float(re.sub(r'[\$,\s\(\)]', '', m1)) if m1 and m1 != "-" else 0
                v2 = float(re.sub(r'[\$,\s\(\)]', '', m2)) if m2 and m2 != "-" else 0
                if not (m1 and m2 and 1990 <= v1 <= 2030 and 1990 <= v2 <= 2030):
                    if m1 or m2: pairs.append((concepto, m1, m2))
            except (ValueError, AttributeError):
                if m1 or m2: pairs.append((concepto, m1, m2))
        else:
            # Impar, sacamos uno por renglón
            for i, c in enumerate(data_cells):
                m1 = _clean_monto_ocr(c["text"].strip())
                if m1:
                    try:
                        v1 = float(re.sub(r'[\$,\s\(\)]', '', m1)) if m1 and m1 != "-" else 0
                        if 1990 <= v1 <= 2030:
                            continue
                    except (ValueError, AttributeError):
                        pass
                    suffix = f" (Dato {i+1})" if len(data_cells) > 1 else ""
                    pairs.append((f"{concepto}{suffix}", m1, ""))
                    
    return pairs


def _extract_pairs_dictaminado(row) -> list:
    """
    Extrae pares (Concepto, Monto1, Monto2) de una fila de dictaminado.

    REGLA CLAVE: Los dictaminados tienen columnas como:
        Concepto | (Nota Ref) | Monto Año 1 | Monto Año 2
    El número de nota (ej. "9") queda en medio de los textos.
    Para no confundirlo con un monto financiero, SIEMPRE tomamos
    los ÚLTIMOS DOS números de la fila como Monto1 y Monto2.

    CASO ESPECIAL: "Nota 6" → el "6" se clasifica como numérico pero
    debe quedar en el concepto para que el match de notas funcione.
    Detectamos esto: si un número es entero pequeño (1-30) y el token
    anterior es "nota" o similar, lo dejamos en el concepto.
    """
    tokens = _tokenize_cells(row)
    if not tokens:
        return []

    # Separar textos y numéricos, PERO preservando números de referencia a notas
    texts = []
    all_numbers = []

    for i, t in enumerate(tokens):
        if t in _OCR_NOISE:
            continue

        if _is_numeric(t) or t == "-":
            # Check if this number is a nota reference (small integer after "nota")
            is_nota_ref = False
            try:
                val = float(t.replace(",", "").replace("$", "").strip())
                # Small integer = likely nota ref, not a financial amount
                if val == int(val) and 1 <= val <= 50:
                    # Look backwards in collected texts for "nota" nearby
                    recent_text = " ".join(texts[-3:]).lower()
                    if "nota" in recent_text or recent_text.strip().endswith("("):
                        is_nota_ref = True
            except (ValueError, AttributeError):
                pass

            if is_nota_ref:
                texts.append(t)  # keep in concept: "INVENTARIOS ( Nota 6 )"
            elif t == "-":
                all_numbers.append(t)
            else:
                all_numbers.append(t)
        else:
            texts.append(t)

    # Los últimos dos son los montos financieros; cualquier número antes es referencia de nota
    if len(all_numbers) >= 2:
        m1 = all_numbers[-2]
        m2 = all_numbers[-1]
    elif len(all_numbers) == 1:
        m1 = all_numbers[-1]
        m2 = ""
    else:
        m1 = ""
        m2 = ""

    concepto = " ".join(texts).strip()

    return [(concepto, m1, m2)]


def _write_nota_data_row(ws, row_num: int, concept: str, monto: str, p_num: int, ev_b64, is_last: bool = False):
    """
    Escribe una fila de datos de nota con fondo azul celeste.
    La última fila tiene un azul ligeramente más intenso para marcar el cierre del bloque.
    """
    fill = NOTA_END_FILL if is_last else NOTA_ROW_FILL

    a = ws[f"A{row_num}"]
    a.value = concept
    a.fill = fill
    a.border = THIN
    a.alignment = Alignment(horizontal="left", indent=2)

    b = ws[f"B{row_num}"]
    b.value = monto
    b.fill = fill
    b.alignment = Alignment(horizontal="right")
    b.border = THIN
    if monto and '(' in str(monto) and ')' in str(monto):
        b.font = Font(color="FF0000")

    c = ws[f"C{row_num}"]
    c.value = p_num
    c.fill = fill
    c.alignment = Alignment(horizontal="center")
    c.border = THIN

    ws[f"D{row_num}"].fill = fill
    ws[f"D{row_num}"].border = THIN
    img_height = _write_evidence_image(ws, "D", row_num, ev_b64)

    e = ws[f"E{row_num}"]
    e.fill = INPUT_FILL
    e.alignment = Alignment(horizontal="right", vertical="center")
    e.number_format = '#,##0.00'
    e.border = THIN

    ws.row_dimensions[row_num].height = max(20, img_height)


def _write_nota_image_row(ws, row_num: int, img_b64: str, p_num: int):
    """
    Escribe una fila de imagen para notas que son solo texto (sin tabla).
    Inserta la captura de pantalla de la sección de texto en la columna D
    con fondo azul celeste, haciendo la fila suficientemente alta.
    """
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws[f"{col}{row_num}"]
        cell.fill = NOTA_ROW_FILL
        cell.border = THIN

    ws[f"C{row_num}"].value = p_num
    ws[f"C{row_num}"].alignment = Alignment(horizontal="center")

    img_height_pt = 80  # default tall height
    if img_b64:
        try:
            img_data = base64.b64decode(img_b64)
            pil_img = PILImage.open(io.BytesIO(img_data))
            orig_w, orig_h = pil_img.size
            # Scale to fit within a wide cell (columns A-D = ~400px wide, tall allowed)
            max_w, max_h = 760, 120
            ratio = min(max_w / orig_w, max_h / orig_h)
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)
            pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

            tmp = io.BytesIO()
            pil_img.save(tmp, format="PNG")
            tmp.seek(0)

            xl_img = OpenpyxlImage(tmp)
            xl_img.anchor = f"A{row_num}"
            ws.add_image(xl_img)
            img_height_pt = max(80, new_h * 0.75)
        except Exception as e:
            logger.warning(f"Could not insert nota image at row {row_num}: {e}")

    ws.row_dimensions[row_num].height = img_height_pt


def _write_evidence_image(ws, col: str, row_num: int, ev_b64: str) -> float:
    """Inserta imagen de evidencia en la celda, retorna altura de imagen."""
    img_height = 20
    if not ev_b64:
        return img_height
    try:
        img_data = base64.b64decode(ev_b64)
        pil_img = PILImage.open(io.BytesIO(img_data))
        orig_w, orig_h = pil_img.size
        max_w, max_h = 400, 60
        ratio = min(max_w / orig_w, max_h / orig_h)
        if ratio < 1:
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)
            pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
        else:
            new_w, new_h = orig_w, orig_h

        tmp = io.BytesIO()
        pil_img.save(tmp, format="PNG")
        tmp.seek(0)
        xl_img = OpenpyxlImage(tmp)
        xl_img.width = new_w
        xl_img.height = new_h
        ws.add_image(xl_img, f"{col}{row_num}")
        img_height = max(img_height, new_h * 0.75 + 5)
    except Exception as e:
        logger.error(f"Error insertando imagen de evidencia: {e}")
    return img_height


def _write_nota_header(ws, row_num: int, nota_label: str):
    """Escribe el header de una nota como celda fusionada A+B con color azul."""
    a = ws[f"A{row_num}"]
    a.value = nota_label
    a.font = NOTA_FONT
    a.fill = NOTA_FILL
    a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    a.border = THIN

    b = ws[f"B{row_num}"]
    b.fill = NOTA_FILL
    b.border = THIN

    ws.merge_cells(f"A{row_num}:B{row_num}")
    ws.row_dimensions[row_num].height = 18


def _write_data_row(ws, row_num: int, concept: str, monto: str, p_num: int, ev_b64):
    """Escribe una fila de datos en las columnas A-E."""
    ws[f"A{row_num}"].value = concept
    ws[f"A{row_num}"].border = THIN

    b_cell = ws[f"B{row_num}"]
    b_cell.value = monto
    b_cell.alignment = Alignment(horizontal="right")
    b_cell.border = THIN
    if monto and '(' in str(monto) and ')' in str(monto):
        b_cell.font = Font(color="FF0000")

    c_cell = ws[f"C{row_num}"]
    c_cell.value = p_num
    c_cell.alignment = Alignment(horizontal="center")
    c_cell.border = THIN

    ws[f"D{row_num}"].border = THIN
    img_height = _write_evidence_image(ws, "D", row_num, ev_b64)

    e = ws[f"E{row_num}"]
    e.fill = INPUT_FILL
    e.alignment = Alignment(horizontal="right", vertical="center")
    e.number_format = '#,##0.00'
    e.border = THIN

    ws.row_dimensions[row_num].height = max(20, img_height)


# ── Main public function ──────────────────────────────────────────────────────

def inject_dictaminado_sheets(doc, wb, mapa):
    """
    Inyecta DOS hojas en el workbook, una por cada año detectado en el dictaminado.
    Soporta tanto el modo normal como el modo notas_dictaminado.
    """
    year_str = str(doc.get("year", "")).strip()

    if "," in year_str:
        years = [y.strip() for y in year_str.split(",") if y.strip()]
    else:
        years = [year_str]

    if not years or not any(years) or years[0] == "Desconocido":
        m = re.search(r'\b(20[1-2]\d)\b', doc.get("filename", ""))
        if m:
            years = [m.group(1)]
        else:
            years = [f"Desc_{uuid.uuid4().hex[:4]}"]

    for year_idx, current_year in enumerate(years):
        sheet_name = current_year
        orig = sheet_name
        cnt = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{orig}_{cnt}"
            cnt += 1

        ws = wb.create_sheet(title=sheet_name)

        # ── Column headers A-E + G-H ──────────────────────────────────────────
        col_headers = {
            "A": ("Cuenta Extraída", 32),
            "B": ("Monto Extraído", 18),
            "C": ("Página", 8),
            "D": ("Evidencia Visual", 55),
            "E": ("Input / Ajuste", 18),
        }
        for col, (title, width) in col_headers.items():
            cell = ws[f"{col}1"]
            cell.value = title
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN
            ws.column_dimensions[col].width = width

        ws.column_dimensions["F"].width = 3
        for col, (title, width) in {"G": ("Concepto", 32), "H": ("Importe (Input)", 18)}.items():
            cell = ws[f"{col}1"]
            cell.value = title
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN
            ws.column_dimensions[col].width = width

        # ── Collect Extracted Data Rows ─────────────────────────────────────────
        main_rows = []
        nota_tables = {} # key: nota number string, value: list of row tuples
        
        if "extracted_data" in doc and "pages" in doc["extracted_data"]:
            for page_data in doc["extracted_data"]["pages"]:
                p_num = page_data.get("page_num", 0) + 1
                layout_type = page_data.get("layout_type", "dictaminado")

                # Track current nota table we are building
                current_nota_num = None
                
                for table in page_data.get("tables", []):
                    for row in table:
                        if not row:
                            continue

                        # Check if first cell is a nota header (injected by extractor)
                        if row and row[0].get("is_nota_header"):
                            nota_label = row[0].get("text", "NOTA")
                            # Extract nota number
                            m = re.search(r'NOTA\s*(\d+)', nota_label, re.IGNORECASE)
                            if m:
                                current_nota_num = m.group(1)
                                if current_nota_num not in nota_tables:
                                    nota_tables[current_nota_num] = []
                            # Only add the header ONCE per nota number to avoid duplicates
                            if current_nota_num:
                                already_has_header = any(
                                    item[0] == "__NOTA_HEADER__"
                                    for item in nota_tables[current_nota_num]
                                )
                                if not already_has_header:
                                    nota_tables[current_nota_num].append(("__NOTA_HEADER__", nota_label, None, None, None))
                            continue

                        # Check if this is a text-only nota screenshot row
                        if row and row[0].get("is_nota_image"):
                            img_b64 = row[0].get("image_b64")
                            if current_nota_num and img_b64:
                                nota_tables[current_nota_num].append(("__NOTA_IMAGE__", img_b64, p_num, None, None))
                            continue

                        # Evidence from any cell in this row
                        evidence_b64 = None
                        for cell_data in row:
                            if cell_data and cell_data.get("evidence_b64"):
                                evidence_b64 = cell_data["evidence_b64"]
                                break

                        # Extract Concept + amounts
                        if layout_type == "notas_dictaminado":
                            d_pairs = _extract_pairs_from_native_cells(row)
                        else:
                            d_pairs = _extract_pairs_dictaminado(row)
                            
                        for concept, m1, m2 in d_pairs:
                            monto = m1 if year_idx == 0 else m2
                            if not concept and not monto:
                                continue
                                
                            row_data = (concept, monto, p_num, evidence_b64)
                            if current_nota_num:
                                nota_tables[current_nota_num].append(("__DATA__",) + row_data)
                            else:
                                main_rows.append(row_data)
                    
                    # Reset current nota table after finishing a table
                    current_nota_num = None

        # ── Interleave and Write Data Rows ────────────────────────────────────
        data_row = 2
        used_notas = set()

        for (concept, monto, p_num, evidence_b64) in main_rows:
            # Write the main row
            _write_data_row(ws, data_row, concept, monto, p_num, evidence_b64)
            data_row += 1
            
            # Check if this concept references a nota
            if concept:
                m = re.search(r'\bnota\s*(\d+)\b', concept, re.IGNORECASE)
                if m:
                    ref_num = m.group(1)
                    if ref_num in nota_tables and ref_num not in used_notas:
                        used_notas.add(ref_num)
                        # Insert the nota table immediately below, with light blue styling
                        items = nota_tables[ref_num]
                        for idx, item in enumerate(items):
                            if item[0] == "__NOTA_HEADER__":
                                _write_nota_header(ws, data_row, item[1])
                                data_row += 1
                            elif item[0] == "__NOTA_IMAGE__":
                                _write_nota_image_row(ws, data_row, item[1], item[2])
                                data_row += 1
                            else:
                                is_last = (idx == len(items) - 1)
                                _write_nota_data_row(ws, data_row, item[1], item[2], item[3], item[4], is_last)
                                data_row += 1
                                
        # Write any unused notas at the bottom
        for nota_num, items in nota_tables.items():
            if nota_num not in used_notas:
                for idx, item in enumerate(items):
                    if item[0] == "__NOTA_HEADER__":
                        _write_nota_header(ws, data_row, item[1])
                        data_row += 1
                    elif item[0] == "__NOTA_IMAGE__":
                        _write_nota_image_row(ws, data_row, item[1], item[2])
                        data_row += 1
                    else:
                        is_last = (idx == len(items) - 1)
                        _write_nota_data_row(ws, data_row, item[1], item[2], item[3], item[4], is_last)
                        data_row += 1

        # ── Structured map (G-H columns) ──────────────────────────────────────
        input_row = 2
        section_rows = {}
        
        # Flatten all extracted data for fuzzy matching
        flat_data = []
        for (concept, monto, p_num, ev) in main_rows:
            if concept and monto:
                flat_data.append((concept, monto))
        for nota_num, items in nota_tables.items():
            for item in items:
                if item[0] == "__DATA__":
                    concept = item[1]
                    monto = item[2]
                    if concept and monto:
                        flat_data.append((concept, monto))

        for tpl_sheet in ["Balance", "Edo de resultados"]:
            if tpl_sheet not in mapa or current_year not in mapa[tpl_sheet]:
                continue

            concepts = mapa[tpl_sheet][current_year]

            hdr = ws[f"G{input_row}"]
            hdr.value = f"── {tpl_sheet.upper()} ──"
            hdr.font = Font(bold=True, color="FFFFFF", size=11)
            hdr.fill = HEADER_FILL
            hdr.alignment = Alignment(horizontal="center", vertical="center")
            hdr.border = THIN
            ws[f"H{input_row}"].fill = HEADER_FILL
            ws[f"H{input_row}"].border = THIN
            ws.merge_cells(f"G{input_row}:H{input_row}")
            input_row += 1

            current_section = None
            current_section_header_row = None
            current_section_first_item = None

            for concept_name, target_cell in concepts.items():
                row_match = re.search(r'\d+', target_cell)
                if not row_match:
                    continue
                tpl_row = int(row_match.group())

                sec_label, sec_color = _get_section(tpl_row, tpl_sheet)
                if sec_label and sec_label != current_section:
                    # Save previous section range
                    if current_section and current_section_first_item:
                        section_rows[current_section] = {
                            "header_row": current_section_header_row,
                            "first_item": current_section_first_item,
                            "last_item": input_row - 1
                        }

                    g = ws[f"G{input_row}"]
                    g.value = sec_label
                    g.font = Font(bold=True, color="FFFFFF", size=10)
                    g.fill = PatternFill("solid", fgColor=sec_color)
                    g.alignment = Alignment(horizontal="left", vertical="center")
                    g.border = THIN
                    ws[f"H{input_row}"].fill = PatternFill("solid", fgColor=sec_color)
                    ws[f"H{input_row}"].border = THIN
                    current_section_header_row = input_row
                    input_row += 1
                    current_section = sec_label
                    current_section_first_item = input_row

                # Nombre del concepto
                c_cell = ws[f"G{input_row}"]
                display_concept = concept_name.replace("_", " ").title()
                c_cell.value = display_concept
                c_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                c_cell.border = THIN

                val_cell = ws[f"H{input_row}"]
                val_cell.fill = INPUT_FILL
                val_cell.number_format = '#,##0.00'
                val_cell.border = THIN
                val_cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Auto-fill using fuzzy match (threshold 0.85 to avoid 'deudores' matching 'acreedores' which is 0.833)
                best_monto_str = _find_best_match(display_concept, flat_data, threshold=0.85)
                if best_monto_str:
                    val_cell.value = _parse_monto(best_monto_str)
                
                # Inyectar fórmula en la plantilla para que apunte a esta hoja de dictaminado
                if tpl_sheet in wb.sheetnames and target_cell:
                    wb[tpl_sheet][target_cell] = f"='{sheet_name}'!H{input_row}"
                
                input_row += 1

            # Save last section
            if current_section and current_section_first_item:
                section_rows[current_section] = {
                    "header_row": current_section_header_row,
                    "first_item": current_section_first_item,
                    "last_item": input_row - 1
                }

            ws[f"G{input_row}"].border = THIN
            ws[f"H{input_row}"].border = THIN
            input_row += 1

        # ── Formulas + verification block ─────────────────────────────────────
        SUM_FONT          = Font(bold=True, color="FFFFFF", size=10)
        COMPROBACION_FILL = PatternFill("solid", fgColor="1565C0")
        COMPROBACION_FONT = Font(bold=True, color="FFFFFF", size=11)
        RESULT_FILL       = PatternFill("solid", fgColor="E8F5E9")
        RESULT_FONT       = Font(bold=True, color="1B5E20", size=10)

        for sec_name, sec_info in section_rows.items():
            h_cell = ws[f"H{sec_info['header_row']}"]
            h_cell.value = f"=SUM(H{sec_info['first_item']}:H{sec_info['last_item']})"
            h_cell.font = SUM_FONT
            h_cell.number_format = '#,##0.00'
            h_cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.column_dimensions["I"].width = 3
        ws.column_dimensions["J"].width = 26
        ws.column_dimensions["K"].width = 20

        j1 = ws["J1"]
        j1.value = "COMPROBACION"
        j1.font = COMPROBACION_FONT
        j1.fill = COMPROBACION_FILL
        j1.alignment = Alignment(horizontal="center", vertical="center")
        j1.border = THIN
        ws["K1"].fill = COMPROBACION_FILL
        ws["K1"].border = THIN
        ws.merge_cells("J1:K1")

        ac_row  = section_rows.get("ACTIVO CIRCULANTE",  {}).get("header_row")
        af_row  = section_rows.get("ACTIVO FIJO",         {}).get("header_row")
        ad_row  = section_rows.get("ACTIVO DIFERIDO",     {}).get("header_row")
        pc_row  = section_rows.get("PASIVO CIRCULANTE",   {}).get("header_row")
        plp_row = section_rows.get("PASIVO LARGO PLAZO",  {}).get("header_row")
        cc_row  = section_rows.get("CAPITAL CONTABLE",    {}).get("header_row")

        verification = [
            ("Total Activos",            f"=H{ac_row}+H{af_row}+H{ad_row}" if ac_row and af_row and ad_row else ""),
            ("Total Pasivos",            f"=H{pc_row}+H{plp_row}"           if pc_row and plp_row         else ""),
            ("Capital Contable",         f"=H{cc_row}"                       if cc_row                     else ""),
            ("Activo-(Pasivo+Capital)", "=K2-(K3+K4)"),
        ]

        for i, (label, formula) in enumerate(verification, start=2):
            j = ws[f"J{i}"]
            j.value = label
            j.font = Font(bold=True, size=10)
            j.alignment = Alignment(horizontal="left", vertical="center")
            j.border = THIN
            j.fill = RESULT_FILL

            k = ws[f"K{i}"]
            k.value = formula
            k.font = RESULT_FONT
            k.number_format = '#,##0.00'
            k.alignment = Alignment(horizontal="right", vertical="center")
            k.border = THIN
            k.fill = RESULT_FILL

        ws["J6"].value = "Resultado:"
        ws["J6"].font = Font(bold=True, size=10)
        ws["J6"].alignment = Alignment(horizontal="left", vertical="center")
        ws["J6"].border = THIN
        ws["K6"].value = '=IF(ABS(K5)<0.01,"SI CUADRA","NO CUADRA")'
        ws["K6"].font = Font(bold=True, size=11)
        ws["K6"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K6"].border = THIN

    return wb
