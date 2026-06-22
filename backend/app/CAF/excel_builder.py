import io
import json
import logging
import re
import fitz
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path("templates/CAF - BRIGHTEC - 2026.02 - plantilla-balance (1).xlsx")
MAPA_PATH = Path("templates/mapa.json")

THIN = Border(
    top=Side("thin", "000000"), left=Side("thin", "000000"),
    right=Side("thin", "000000"), bottom=Side("thin", "000000"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

# Secciones del Balance según fila en la plantilla
BALANCE_SECTIONS = [
    (range(9, 19),  "ACTIVO CIRCULANTE",  "1A7A4A"),
    (range(24, 32), "ACTIVO FIJO",         "2E7D32"),
    (range(37, 40), "ACTIVO DIFERIDO",     "388E3C"),
    (range(52, 58), "PASIVO CIRCULANTE",   "C62828"),
    (range(64, 66), "PASIVO LARGO PLAZO",  "E53935"),
    (range(74, 79), "CAPITAL CONTABLE",    "6A1B9A"),
]
EDO_SECTIONS = [
    (range(7, 9),   "INGRESOS",            "1A7A4A"),
    (range(8, 12),  "COSTOS Y GASTOS",     "C62828"),
    (range(15, 18), "RESULTADO FINANCIERO","1565C0"),
    (range(20, 22), "OTROS",               "6D4C41"),
    (range(24, 25), "IMPUESTOS",           "4E342E"),
    (range(29, 30), "DEPRECIACIÓN",        "37474F"),
]

def _get_section(row_num, sheet_name):
    for rng, label, color in (BALANCE_SECTIONS if sheet_name == "Balance" else EDO_SECTIONS):
        if row_num in rng:
            return label, color
    return None, None


def _is_numeric(text):
    """Checa si un texto parece un número (con comas, $, negativos, paréntesis, etc.)."""
    cleaned = text.replace(",", "").replace("$", "").replace(" ", "").replace("-", "").replace(".", "").replace("(", "").replace(")", "")
    return cleaned.isdigit() and len(cleaned) > 0


# Ruido de OCR que se debe ignorar
_OCR_NOISE = {"$", "S", "EA", "69", "SSS", "EAEAEA", "6969", "6A", "09", "th"}


def _tokenize_cells(cells):
    """Extrae tokens leyendo las celdas línea por línea horizontalmente."""
    tokens = []
    
    # Extraer las líneas de cada celda
    lines_per_cell = []
    for c in cells:
        if not c or not str(c.get("text", "")).strip():
            lines_per_cell.append([])
            continue
        # Separar por saltos de línea y limpiar
        cell_lines = [line.strip() for line in str(c["text"]).strip().split('\n')]
        lines_per_cell.append(cell_lines)
        
    if not lines_per_cell:
        return []
        
    # Encontrar la celda con más líneas
    max_lines = max(len(lines) for lines in lines_per_cell)
    
    # Leer línea por línea de izquierda a derecha (transponer)
    for i in range(max_lines):
        for cell_lines in lines_per_cell:
            if i < len(cell_lines) and cell_lines[i]:
                tokens.append(cell_lines[i])
                
    return tokens


def _pair_tokens(tokens):
    """Agrupa tokens en pares (concepto, monto) de forma lineal."""
    pairs = []
    current_concept_parts = []

    for t in tokens:
        if t in _OCR_NOISE:
            continue
        if _is_numeric(t) or t == "-":
            concepto = " ".join(current_concept_parts).strip()
            pairs.append((concepto, t))
            current_concept_parts = []
        else:
            current_concept_parts.append(t)

    leftover = " ".join(current_concept_parts).strip()
    if leftover:
        pairs.append((leftover, ""))

    if not pairs and tokens:
        pairs.append((" | ".join(tokens), ""))

    return pairs


def _extract_pairs_single_column(row):
    """Estrategia LINEAL: tokeniza toda la fila y empareja concepto→monto secuencialmente."""
    tokens = _tokenize_cells(row)
    if not tokens:
        return []
        
    # Si despues de tokenizar solo hay 1 token, intentamos separarlo por Regex
    # Porque Document AI o el fallback lo metió todo en una sola línea/celda
    if len(tokens) == 1:
        import re
        t = tokens[0]
        m1 = re.match(r'^([\$0-9,\.\-\(\)]+)\s+(.+)$', t)
        m2 = re.match(r'^(.+)\s+([\$0-9,\.\-\(\)]+)$', t)
        if m1 and _is_numeric(m1.group(1)):
            tokens = [m1.group(1), m1.group(2)]
        elif m2 and _is_numeric(m2.group(2)):
            tokens = [m2.group(1), m2.group(2)]
            
    return _pair_tokens(tokens)


def _extract_pairs_dictaminado(row):
    tokens = _tokenize_cells(row)
    if not tokens: return []
    pairs = []
    current_concept = []
    amounts = []
    for t in tokens:
        if t in _OCR_NOISE: continue
        if _is_numeric(t) or t == "-":
            amounts.append(t)
        else:
            if amounts:
                c = " ".join(current_concept).strip()
                m1 = amounts[0] if len(amounts) > 0 else ""
                m2 = amounts[1] if len(amounts) > 1 else ""
                pairs.append((c, m1, m2))
                current_concept = [t]
                amounts = []
            else:
                current_concept.append(t)
    if current_concept or amounts:
        c = " ".join(current_concept).strip()
        m1 = amounts[0] if len(amounts) > 0 else ""
        m2 = amounts[1] if len(amounts) > 1 else ""
        pairs.append((c, m1, m2))
    return pairs

def _extract_pairs_two_column(row, page_width, regions=None):
    """
    Since extractor.py physically crops the images, ALL cells in `row` 
    belong to the SAME column. We just route the entire row's pairs to 
    left or right based on its x-coordinate.
    """
    first_cell = next((c for c in row if c and c.get("bbox")), None)
    if not first_cell:
        return [], []
        
    x0 = first_cell["bbox"][0]
    split_line = page_width / 2.0
    if regions and len(regions) >= 2:
        sorted_regions = sorted(regions, key=lambda r: r["x"])
        # Same split logic as extractor.py
        r1 = sorted_regions[0]
        r2 = sorted_regions[1]
        r1_end = r1["x"] + r1["w"]
        r2_start = r2["x"]
        split_x_norm = r2_start - 0.015
        if split_x_norm <= r1_end:
            split_x_norm = (r1_end + r2_start) / 2.0
        split_line = split_x_norm * page_width

    # Parse the row using the robust single_column logic
    pairs = _extract_pairs_single_column(row)
    
    if x0 < split_line:
        return pairs, []
    else:
        return [], pairs


def build_caf_excel(docs_data: list) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {TEMPLATE_PATH}")
    if not MAPA_PATH.exists():
        raise FileNotFoundError(f"Mapa no encontrado: {MAPA_PATH}")

    with open(MAPA_PATH, "r", encoding="utf-8") as f:
        mapa = json.load(f)

    wb = load_workbook(TEMPLATE_PATH)

    for doc in docs_data:
        doc_type = doc.get("doc_type", "financiero")
        if "extracted_data" in doc and "doc_type" in doc["extracted_data"]:
            doc_type = doc["extracted_data"]["doc_type"]
            
        is_dictaminado = (doc_type == "dictaminado")
        year_str = str(doc.get("year", "")).strip()
        
        if is_dictaminado and "," in year_str:
            years = [y.strip() for y in year_str.split(",") if y.strip()]
        else:
            years = [year_str]
            
        if not years or not any(years) or years[0] == "Desconocido":
            m = re.search(r'\b(20[1-2]\d)\b', doc.get("filename", ""))
            if m:
                years = [m.group(1)]
            else:
                import uuid
                years = [f"Desc_{uuid.uuid4().hex[:4]}"]

        for year_idx, current_year in enumerate(years):
            sheet_name = current_year
            orig = sheet_name
            cnt = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{orig}_{cnt}"
                cnt += 1

            ws = wb.create_sheet(title=sheet_name)

            # ══════════════════════════════════════════════════════════
            # HEADERS
            # ══════════════════════════════════════════════════════════
            headers = {
                "A": ("Cuenta Extraída", 28),
                "B": ("Monto Extraído", 18),
                "C": ("Página", 8),
                "D": ("Evidencia Visual", 55),
                "E": ("Input / Ajuste", 18),
            }
            for col, (title, width) in headers.items():
                cell = ws[f"{col}1"]
                cell.value = title
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN
                ws.column_dimensions[col].width = width

            # Separador + Headers del mapa estructurado (más a la derecha)
            ws.column_dimensions["F"].width = 3  # separador
            for col, (title, width) in {"G": ("Concepto", 32), "H": ("Importe (Input)", 18)}.items():
                cell = ws[f"{col}1"]
                cell.value = title
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN
                ws.column_dimensions[col].width = width

            # ══════════════════════════════════════════════════════════
            # COLUMNAS A-E: Datos extraídos del PDF (una fila por dato)
            # ══════════════════════════════════════════════════════════
            page_layouts = doc.get("page_layouts", {})
            if not page_layouts and "extracted_data" in doc:
                page_layouts = doc["extracted_data"].get("page_layouts", {})

            data_row = 2
            if "extracted_data" in doc and "pages" in doc["extracted_data"]:
                for page_data in doc["extracted_data"]["pages"]:
                    p_index = page_data.get("page_num", 0)
                    p_num = p_index + 1
                    page_width = page_data.get("page_width", 600)
                    layout_val = page_layouts.get(str(p_index), "single_column")
                    layout_type = layout_val
                    regions = None
                    if isinstance(layout_val, dict):
                        layout_type = layout_val.get("type", "single_column")
                        regions = layout_val.get("regions", None)

                    page_left_pairs = []
                    page_right_pairs = []
                    page_single_pairs = []

                    for table in page_data.get("tables", []):
                        for row in table:
                            if not row:
                                continue

                            # ── Extraer la imagen de evidencia de la fila completa ──
                            evidence_b64 = None
                            for cell_data in row:
                                if cell_data and cell_data.get("evidence_b64"):
                                    evidence_b64 = cell_data["evidence_b64"]
                                    break

                            # ── Decidir estrategia según layout ──
                            if layout_type == "two_column":
                                l_pairs, r_pairs = _extract_pairs_two_column(row, page_width, regions)
                                for c, m in l_pairs: page_left_pairs.append((c, m, evidence_b64))
                                for c, m in r_pairs: page_right_pairs.append((c, m, evidence_b64))
                            elif layout_type == "split_column":
                                concept_text = ""
                                amount_text = ""
                                for cell in row:
                                    if cell.get("is_concept"): concept_text = cell.get("text", "")
                                    elif cell.get("is_amount"): amount_text = cell.get("text", "")
                                page_single_pairs.append((concept_text, amount_text, evidence_b64))
                            elif is_dictaminado:
                                d_pairs = _extract_pairs_dictaminado(row)
                                for c, m1, m2 in d_pairs:
                                    m = m1 if year_idx == 0 else m2
                                    page_single_pairs.append((c, m, evidence_b64))
                            else:
                                # single_column o auto → lineal
                                pairs = _extract_pairs_single_column(row)
                                for c, m in pairs: page_single_pairs.append((c, m, evidence_b64))

                    # Unir orfandades antes de escribir (si el OCR separó conceptos y montos en filas distintas)
                    def _cleanup_orphan_pairs(pairs_list):
                        cleaned = []
                        pending_concepts = []
                        for c, m, ev in pairs_list:
                            if c and not m:
                                pending_concepts.append((c, ev))
                            elif not c and m:
                                if pending_concepts:
                                    pc, pev = pending_concepts.pop(0) # FIFO
                                    cleaned.append((pc, m, pev or ev))
                                else:
                                    cleaned.append((c, m, ev))
                            else:
                                for pc, pev in pending_concepts:
                                    cleaned.append((pc, "", pev))
                                pending_concepts = []
                                cleaned.append((c, m, ev))
                        for pc, pev in pending_concepts:
                            cleaned.append((pc, "", pev))
                        return cleaned

                    # Agrupar los pares para escribirlos: si es doble columna, primero la izquierda, luego la derecha
                    if layout_type == "two_column":
                        all_page_pairs = _cleanup_orphan_pairs(page_left_pairs) + _cleanup_orphan_pairs(page_right_pairs)
                    else:
                        all_page_pairs = _cleanup_orphan_pairs(page_single_pairs)

                    # ── Escribir cada par como fila en Excel ──
                    for concepto, monto, evidence_b64 in all_page_pairs:
                                if not concepto and not monto:
                                    continue
                                if not concepto and monto:
                                    concepto = "(Monto sin concepto)"

                                # Col A: Cuenta
                                a = ws[f"A{data_row}"]
                                a.value = concepto
                                a.alignment = Alignment(vertical="center", wrap_text=True)
                                a.border = THIN

                                # Col B: Monto
                                b = ws[f"B{data_row}"]
                                b.value = monto
                                b.alignment = Alignment(vertical="center", horizontal="right")
                                b.border = THIN

                                # Col C: Página
                                c = ws[f"C{data_row}"]
                                c.value = f"Pág {p_num}"
                                c.alignment = Alignment(vertical="center", horizontal="center")
                                c.border = THIN

                                # Col D: Imagen recortada
                                img_height = 30
                                if evidence_b64:
                                    try:
                                        import base64
                                        img_data = base64.b64decode(evidence_b64)
                                        pil_img = PILImage.open(io.BytesIO(img_data))
                                        w, h = pil_img.size

                                        max_w = 400
                                        if w > max_w:
                                            ratio = max_w / w
                                            w, h = max_w, int(h * ratio)

                                        xl_img = OpenpyxlImage(io.BytesIO(img_data))
                                        xl_img.width = w
                                        xl_img.height = h
                                        ws.add_image(xl_img, f"D{data_row}")
                                        img_height = h * 0.75 + 5
                                    except Exception as e:
                                        logger.error(f"Error insertando imagen de evidencia: {e}")

                                # Col E: Input / Ajuste
                                e = ws[f"E{data_row}"]
                                e.fill = INPUT_FILL
                                e.alignment = Alignment(horizontal="right", vertical="center")
                                e.number_format = '#,##0.00'
                                e.border = THIN

                                ws.row_dimensions[data_row].height = max(20, img_height)
                                data_row += 1

            # ══════════════════════════════════════════════════════════
            # COLUMNAS G-H: Inputs estructurados del mapa.json
            # (enlazados a la plantilla de Balance / Edo de resultados)
            # ══════════════════════════════════════════════════════════
            input_row = 2
            section_rows = {}  # Track where each section header and its items are

            for tpl_sheet in ["Balance", "Edo de resultados"]:
                if tpl_sheet not in mapa or current_year not in mapa[tpl_sheet]:
                    continue

                concepts = mapa[tpl_sheet][current_year]

                # Encabezado de sección principal
                hdr = ws[f"G{input_row}"]
                hdr.value = f"── {tpl_sheet.upper()} ──"
                hdr.font = Font(bold=True, color="FFFFFF", size=11)
                hdr.fill = HEADER_FILL
                hdr.alignment = Alignment(horizontal="center", vertical="center")
                hdr.border = THIN
                ws[f"H{input_row}"].fill = HEADER_FILL
                ws[f"H{input_row}"].border = THIN
                ws.merge_cells(f"G{input_row}:H{input_row}")
                input_row += 1

                current_section = None
                current_section_header_row = None
                current_section_first_item = None

                for concept_name, target_cell in concepts.items():
                    row_match = re.search(r'\d+', target_cell)
                    if not row_match:
                        continue
                    tpl_row = int(row_match.group())

                    sec_label, sec_color = _get_section(tpl_row, tpl_sheet)
                    if sec_label and sec_label != current_section:
                        # Save previous section range
                        if current_section and current_section_first_item:
                            section_rows[current_section] = {
                                "header_row": current_section_header_row,
                                "first_item": current_section_first_item,
                                "last_item": input_row - 1
                            }

                        g = ws[f"G{input_row}"]
                        g.value = sec_label
                        g.font = Font(bold=True, color="FFFFFF", size=10)
                        g.fill = PatternFill("solid", fgColor=sec_color)
                        g.alignment = Alignment(horizontal="left", vertical="center")
                        g.border = THIN
                        ws[f"H{input_row}"].fill = PatternFill("solid", fgColor=sec_color)
                        ws[f"H{input_row}"].border = THIN
                        current_section_header_row = input_row
                        input_row += 1
                        current_section = sec_label
                        current_section_first_item = input_row

                    # Nombre del concepto
                    g = ws[f"G{input_row}"]
                    g.value = concept_name.replace("_", " ").title()
                    g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    g.border = THIN

                    # Input amarillo
                    h = ws[f"H{input_row}"]
                    h.fill = INPUT_FILL
                    h.alignment = Alignment(horizontal="right", vertical="center")
                    h.number_format = '#,##0.00'
                    h.border = THIN

                    # Fórmula en la plantilla
                    if tpl_sheet in wb.sheetnames and target_cell:
                        wb[tpl_sheet][target_cell] = f"='{sheet_name}'!H{input_row}"

                    input_row += 1

                # Save last section of this sheet
                if current_section and current_section_first_item:
                    section_rows[current_section] = {
                        "header_row": current_section_header_row,
                        "first_item": current_section_first_item,
                        "last_item": input_row - 1
                    }

                input_row += 1

            # ══════════════════════════════════════════════════════════
            # INYECTAR FÓRMULAS DE SUMA EN HEADERS DE SECCIÓN (Col H)
            # ══════════════════════════════════════════════════════════
            SUM_FONT = Font(bold=True, color="FFFFFF", size=10)
            COMPROBACION_FILL = PatternFill("solid", fgColor="1565C0")
            COMPROBACION_FONT = Font(bold=True, color="FFFFFF", size=11)
            RESULT_FILL = PatternFill("solid", fgColor="E8F5E9")
            RESULT_FONT = Font(bold=True, color="1B5E20", size=10)

            for sec_name, sec_info in section_rows.items():
                hr = sec_info["header_row"]
                fi = sec_info["first_item"]
                li = sec_info["last_item"]
                h_cell = ws[f"H{hr}"]
                h_cell.value = f"=SUM(H{fi}:H{li})"
                h_cell.font = SUM_FONT
                h_cell.number_format = '#,##0.00'
                h_cell.alignment = Alignment(horizontal="right", vertical="center")

            # ══════════════════════════════════════════════════════════
            # COLUMNA J-K: BLOQUE DE COMPROBACIÓN CONTABLE
            # ══════════════════════════════════════════════════════════
            ws.column_dimensions["I"].width = 3  # separador
            ws.column_dimensions["J"].width = 26
            ws.column_dimensions["K"].width = 20

            # J1: Header "COMPROBACIÓN"
            j1 = ws["J1"]
            j1.value = "COMPROBACION"
            j1.font = COMPROBACION_FONT
            j1.fill = COMPROBACION_FILL
            j1.alignment = Alignment(horizontal="center", vertical="center")
            j1.border = THIN
            k1 = ws["K1"]
            k1.fill = COMPROBACION_FILL
            k1.border = THIN
            ws.merge_cells("J1:K1")

            ac_row = section_rows.get("ACTIVO CIRCULANTE", {}).get("header_row")
            af_row = section_rows.get("ACTIVO FIJO", {}).get("header_row")
            ad_row = section_rows.get("ACTIVO DIFERIDO", {}).get("header_row")
            pc_row = section_rows.get("PASIVO CIRCULANTE", {}).get("header_row")
            plp_row = section_rows.get("PASIVO LARGO PLAZO", {}).get("header_row")
            cc_row = section_rows.get("CAPITAL CONTABLE", {}).get("header_row")

            verification_rows = [
                ("Total Activos", f"=H{ac_row}+H{af_row}+H{ad_row}" if ac_row and af_row and ad_row else ""),
                ("Total Pasivos", f"=H{pc_row}+H{plp_row}" if pc_row and plp_row else ""),
                ("Capital Contable", f"=H{cc_row}" if cc_row else ""),
                ("Activo-(Pasivo+Capital)", "=K2-(K3+K4)"),
            ]

            for i, (label, formula) in enumerate(verification_rows, start=2):
                j = ws[f"J{i}"]
                j.value = label
                j.font = Font(bold=True, size=10)
                j.alignment = Alignment(horizontal="left", vertical="center")
                j.border = THIN
                j.fill = RESULT_FILL

                k = ws[f"K{i}"]
                k.value = formula
                k.font = RESULT_FONT
                k.number_format = '#,##0.00'
                k.alignment = Alignment(horizontal="right", vertical="center")
                k.border = THIN
                k.fill = RESULT_FILL

            # Fila 6: Indicador visual de si cuadra o no
            ws["J6"].value = "Resultado:"
            ws["J6"].font = Font(bold=True, size=10)
            ws["J6"].alignment = Alignment(horizontal="left", vertical="center")
            ws["J6"].border = THIN
            ws["K6"].value = '=IF(ABS(K5)<0.01,"SI CUADRA","NO CUADRA")'
            ws["K6"].font = Font(bold=True, size=11)
            ws["K6"].alignment = Alignment(horizontal="center", vertical="center")
            ws["K6"].border = THIN

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

