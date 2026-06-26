"""
AutoTeaser Backend - Bank Statement Processor
Main FastAPI application.
"""
import uuid
import logging
from pathlib import Path
from datetime import datetime



from pydantic import BaseModel
from typing import List, Optional

class DocIdsRequest(BaseModel):
    doc_ids: List[str]

class GenerateConsolidatedRequest(BaseModel):
    doc_ids: List[str]
    template_name: str

from app.CAF.routes import router as caf_router
from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from app.config import UPLOAD_DIR, OUTPUT_DIR, TEMPLATES_DIR, MAX_FILE_SIZE
from app.pdf_extractor import extract_text
from app.bank_detector import detect_bank
from app.banks import get_parser
from app.excel_processor import fill_template, read_template_fields



# Portal de Expedientes Imports
from contextlib import asynccontextmanager
from portal.shared.supabase_db import init_supabase
from portal.Cliente import auth as portal_auth
from portal.Cliente import upload as portal_upload
from portal.Cliente import expedientes as portal_expedientes
from portal.Admin import dashboard as portal_dashboard
from portal.Admin import revision as portal_revision

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Inicializa clientes externos al arrancar la aplicación."""
    try:
        init_supabase()
        logger.info("Supabase clients initialized successfully")
    except Exception as e:
        logger.warning(f"Supabase init failed (portal features disabled): {e}")
    yield

# Create app
app = FastAPI(
    title="AutoTeaser - Bank Statement Processor",
    description="Extract data from bank statement PDFs and fill Excel templates",
    version="0.1.0",
    lifespan=lifespan,
)

# CORS - allow frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Portal de Expedientes — Routers
app.include_router(portal_auth.router,         prefix="/api/portal/cliente", tags=["Portal Cliente"])
app.include_router(portal_upload.router,       prefix="/api/portal/cliente", tags=["Portal Cliente"])
app.include_router(portal_expedientes.router,  prefix="/api/portal/cliente", tags=["Portal Cliente"])
app.include_router(portal_dashboard.router,    prefix="/api/portal/admin",   tags=["Portal Admin"])
app.include_router(portal_revision.router,     prefix="/api/portal/admin",   tags=["Portal Admin"])

# In-memory store
documents = {}
caf_documents = {}


# Pydantic models
class BatchFillRequest(BaseModel):
    doc_ids: list[str]
    template_name: str


class MergeTemplatesRequest(BaseModel):
    doc_ids: list[str]
    template_name: Optional[str] = None



# ── Routes ──────────────────────────────────────────────────────────────


@app.get("/")
async def root():
    return {
        "service": "AutoTeaser - Bank Statement Processor",
        "version": "0.1.0",
    }


@app.post("/api/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Upload a bank statement PDF.
    Extracts text immediately and detects the bank.
    """
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, f"Archivo muy grande. Máximo: {MAX_FILE_SIZE // (1024*1024)}MB")

    # Save file
    doc_id = str(uuid.uuid4())[:8]
    safe_name = f"{doc_id}_{file.filename}"
    file_path = UPLOAD_DIR / safe_name

    with open(file_path, "wb") as f:
        f.write(content)

    logger.info(f"PDF uploaded: {safe_name} ({len(content)} bytes)")

    # Extract text
    try:
        extraction = extract_text(file_path)
    except Exception as e:
        logger.error(f"Extraction failed: {e}")
        raise HTTPException(500, f"Error al extraer texto del PDF: {str(e)}")

    # Detect bank (Gemini Vision identifies the logo, text fallback if Gemini unavailable)
    detected_bank = detect_bank(pdf_path=file_path, text=extraction["full_text"])

    # Store
    doc_info = {
        "id": doc_id,
        "file_name": file.filename,
        "file_path": str(file_path),
        "uploaded_at": datetime.now().isoformat(),
        "extraction": extraction,
        "detected_bank": detected_bank,
        "status": "uploaded",
        "parsed_data": None,
        "output_file": None,
    }
    documents[doc_id] = doc_info

    return {
        "id": doc_id,
        "file_name": file.filename,
        "page_count": extraction["page_count"],
        "detected_bank": detected_bank,
        "text_preview": extraction["full_text"][:500],
        "status": "uploaded",
    }


class SetBankRequest(BaseModel):
    bank: str


@app.post("/api/documents/{doc_id}/set-bank")
async def set_document_bank(doc_id: str, request: SetBankRequest):
    """
    Asigna o cambia el banco detectado para un documento subido de forma manual.
    """
    if doc_id not in documents:
        raise HTTPException(404, f"Documento {doc_id} no encontrado")
    
    bank = request.bank.lower().strip()
    from app.bank_detector import VALID_BANKS
    if bank not in VALID_BANKS and bank != "desconocido":
        raise HTTPException(400, f"Banco no válido. Debe ser uno de: {VALID_BANKS}")
        
    documents[doc_id]["detected_bank"] = bank if bank != "desconocido" else None
    return {
        "id": doc_id,
        "detected_bank": documents[doc_id]["detected_bank"],
        "status": documents[doc_id]["status"]
    }


@app.post("/api/process/{doc_id}")
async def process_document(doc_id: str, engine: str = "gemini"):
    """
    Process the uploaded PDF with the bank-specific parser.
    """
    if doc_id not in documents:
        raise HTTPException(404, f"Documento {doc_id} no encontrado")

    doc = documents[doc_id]
    bank = doc["detected_bank"]

    if not bank:
        raise HTTPException(400, "No se pudo detectar el banco. Sube un estado de cuenta válido.")

    parser = get_parser(bank)
    if not parser:
        raise HTTPException(400, f"No hay parser implementado para el banco: {bank}")

    try:
        # Pass pdf_path and engine as kwargs – parsers that need them (HSBC, BBVA)
        # accept **kwargs so others simply ignore them.
        parse_kwargs = {"pdf_path": doc["file_path"]}
        if bank == "hsbc":
            parse_kwargs["engine"] = engine
        
        parsed_data = parser.parse(
            doc["extraction"]["full_text"],
            doc["extraction"]["pages"],
            **parse_kwargs,
        )
        doc["parsed_data"] = parsed_data
        doc["status"] = "processed"
    except NotImplementedError:
        raise HTTPException(501, f"El parser de {bank} aún no está implementado")
    except Exception as e:
        logger.error(f"Parse failed for {bank}: {e}")
        doc["status"] = "error"
        raise HTTPException(500, f"Error al procesar: {str(e)}")

    return {
        "id": doc_id,
        "bank": bank,
        "status": doc["status"],
        "data": parsed_data,
    }


def _load_mapping(template_name: str) -> dict | None:
    import json
    stem = Path(template_name).stem
    possible_names = [
        f"{stem}.json",
        f"{stem.replace(' - ', '-')}.json",
        f"{stem.replace(' ', '')}.json",
        "TEASER-PLANTILLA.json"  # Ultimate fallback for the main template
    ]
    for name in possible_names:
        map_path = TEMPLATES_DIR / name
        if map_path.exists():
            logger.info(f"Loaded mapping file: {map_path.name}")
            with open(map_path, "r", encoding="utf-8") as f:
                return json.load(f)
    logger.warning(f"No mapping JSON found for template {template_name}")
    return None


@app.post("/api/fill-template/{doc_id}")
async def fill_template_endpoint(doc_id: str, template_name: str | None = None):
    """
    Fill an Excel template with parsed data from a single document.
    """
    if doc_id not in documents:
        raise HTTPException(404, f"Documento {doc_id} no encontrado")

    doc = documents[doc_id]

    if not doc["parsed_data"]:
        raise HTTPException(400, "El documento no ha sido procesado. Usa POST /api/process/{doc_id} primero.")

    if not template_name:
        raise HTTPException(400, "Debes especificar una plantilla (template_name)")

    template_path = TEMPLATES_DIR / template_name
    if not template_path.exists():
        raise HTTPException(404, f"Plantilla {template_name} no encontrada")

    mapping = _load_mapping(template_name)

    output_name = f"{doc_id}_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = OUTPUT_DIR / output_name

    try:
        fill_template(
            template_path=str(template_path),
            output_path=str(output_path),
            data_list=[doc["parsed_data"]],
            mapping=mapping,
        )
    except Exception as e:
        logger.error(f"Template fill failed: {e}")
        raise HTTPException(500, f"Error al llenar plantilla: {str(e)}")

    doc["output_file"] = str(output_path)
    doc["status"] = "completed"

    return {
        "id": doc_id,
        "status": "completed",
        "output_file": output_name,
        "download_url": f"/api/download/{doc_id}",
    }


@app.post("/api/fill-template-batch")
async def fill_template_batch_endpoint(req: BatchFillRequest):
    """
    Fill an Excel template with parsed data from multiple documents (batch).
    """
    if not req.doc_ids:
        raise HTTPException(400, "No se especificaron documentos (doc_ids)")

    data_list = []
    for doc_id in req.doc_ids:
        if doc_id not in documents:
            raise HTTPException(404, f"Documento {doc_id} no encontrado")
        doc = documents[doc_id]
        if not doc.get("parsed_data"):
            raise HTTPException(400, f"El documento {doc['file_name']} no ha sido procesado.")
        data_list.append(doc["parsed_data"])

    if not req.template_name or not req.template_name.strip():
        raise HTTPException(400, "No se especificó el nombre de la plantilla (template_name)")

    template_path = TEMPLATES_DIR / req.template_name
    if not template_path.exists() or template_path.is_dir():
        raise HTTPException(404, f"Plantilla {req.template_name} no encontrada")

    mapping = _load_mapping(req.template_name)

    output_name = f"batch_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = OUTPUT_DIR / output_name

    try:
        fill_template(
            template_path=str(template_path),
            output_path=str(output_path),
            data_list=data_list,
            mapping=mapping,
        )
    except Exception as e:
        logger.error(f"Batch template fill failed: {e}")
        raise HTTPException(500, f"Error al llenar plantilla en lote: {str(e)}")

    batch_id = f"batch_{str(uuid.uuid4())[:4]}"
    documents[batch_id] = {
        "id": batch_id,
        "file_name": output_name,
        "detected_bank": "LOTE",
        "status": "completed",
        "uploaded_at": datetime.now().isoformat(),
        "extraction": {"page_count": 0, "full_text": ""},
        "parsed_data": None,
        "output_file": str(output_path),
    }

    return {
        "id": batch_id,
        "status": "completed",
        "output_file": output_name,
        "download_url": f"/api/download/{batch_id}",
    }


@app.post("/api/preview-batch")
async def preview_batch_endpoint(req: BatchFillRequest):
    """
    Generate a preview table of how the data would look in the Excel
    without actually generating the file.
    """
    from collections import defaultdict

    if not req.doc_ids:
        raise HTTPException(400, "No se especificaron documentos (doc_ids)")

    mapping = _load_mapping(req.template_name)
    months_order = mapping.get("months_order", ["nov", "dic", "ene", "feb", "mar", "abr", "may"]) if mapping else []

    # Group data by account
    accounts = defaultdict(dict)
    for doc_id in req.doc_ids:
        if doc_id not in documents:
            continue
        doc = documents[doc_id]
        if not doc.get("parsed_data"):
            continue
        data = doc["parsed_data"]
        acct = data.get("account_name", "").strip()
        month = data.get("month", "").lower().strip()
        if acct and month:
            accounts[acct][month] = {
                "deposits": data.get("deposits", 0.0),
                "average_balance": data.get("average_balance", 0.0),
            }

    # Build table rows
    table = []
    for acct_name in sorted(accounts.keys()):
        months_data = accounts[acct_name]
        row = {"account_name": acct_name, "months": {}}
        for m in months_order:
            if m in months_data:
                row["months"][m] = months_data[m]
            else:
                row["months"][m] = None
        table.append(row)

    return {
        "months_order": months_order,
        "accounts": table,
        "total_accounts": len(table),
    }


@app.get("/api/download/{doc_id}")
async def download_file(doc_id: str):
    """Download the filled Excel."""
    if doc_id not in documents:
        raise HTTPException(404, "Documento no encontrado")

    doc = documents[doc_id]
    if not doc.get("output_file"):
        raise HTTPException(400, "No hay archivo generado aún")

    file_path = Path(doc["output_file"])
    if not file_path.exists():
        raise HTTPException(404, "Archivo no encontrado en disco")

    return FileResponse(
        path=str(file_path),
        filename=f"AutoTeaser_{doc['file_name'].replace('.pdf', '')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/documents")
async def list_documents():
    """List all uploaded documents."""
    docs = []
    for doc in documents.values():
        docs.append({
            "id": doc["id"],
            "file_name": doc["file_name"],
            "detected_bank": doc["detected_bank"],
            "status": doc["status"],
            "uploaded_at": doc["uploaded_at"],
            "page_count": doc["extraction"]["page_count"],
        })
    return {"documents": docs}


@app.get("/api/documents/{doc_id}")
async def get_document(doc_id: str):
    """Get document details."""
    if doc_id not in documents:
        raise HTTPException(404, "Documento no encontrado")

    doc = documents[doc_id]
    return {
        "id": doc["id"],
        "file_name": doc["file_name"],
        "detected_bank": doc["detected_bank"],
        "status": doc["status"],
        "uploaded_at": doc["uploaded_at"],
        "page_count": doc["extraction"]["page_count"],
        "parsed_data": doc["parsed_data"],
        "text_preview": doc["extraction"]["full_text"][:1000],
        "has_output": doc["output_file"] is not None,
        "download_url": f"/api/download/{doc_id}" if doc["output_file"] else None,
    }


@app.delete("/api/documents/{doc_id}")
async def delete_document(doc_id: str):
    """Delete a document and its files."""
    if doc_id not in documents:
        raise HTTPException(404, "Documento no encontrado")

    doc = documents[doc_id]

    if doc.get("file_path") and Path(doc["file_path"]).exists():
        Path(doc["file_path"]).unlink()
    if doc.get("output_file") and Path(doc["output_file"]).exists():
        Path(doc["output_file"]).unlink()

    del documents[doc_id]
    return {"message": f"Documento {doc_id} eliminado"}


@app.get("/api/templates")
async def list_templates():
    """List available Excel templates for bank statements (Teaser)."""
    templates = []
    for f in TEMPLATES_DIR.iterdir():
        if f.suffix in (".xlsx", ".xls"):
            # Check sheet names to filter out CAF templates
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(f), read_only=True)
                sheet_names = [s.lower() for s in wb.sheetnames]
                wb.close()
                is_teaser = any("banco" in s or "cuenta" in s for s in sheet_names)
                if not is_teaser:
                    # Skip files without bank-related sheets
                    continue
            except Exception:
                # If error, fallback to name-based heuristic
                name = f.name.lower()
                if "caf" in name or "balance" in name or "resultados" in name:
                    continue
            
            templates.append({"name": f.name})
    return {"templates": templates}


@app.get("/api/health")
async def health_check():
    return {
        "status": "ok",
        "documents_count": len(documents),
        "caf_documents_count": len(caf_documents),
    }


# ── AutoCAF Routes ──────────────────────────────────────────────────────

# ── AutoCAF v2 ──
app.include_router(caf_router)
