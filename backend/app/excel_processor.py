"""
AutoTeaser - Excel Processor
Uses openpyxl to fill Excel templates with extracted bank statement data.

DYNAMIC BLOCK ASSIGNMENT:
- Groups all parsed data by unique account_name (e.g. "BBVA 0757", "HSBC 1352")
- Assigns each unique account to the next available block in the template
- Within each block, fills months in order by date
- No pre-assigned blocks — everything is determined at runtime
"""
import logging
import shutil
import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Helper to normalize and sort months chronologically
SPANISH_MONTHS = {
    'ene': 1, 'enero': 1, 'jan': 1, 'january': 1,
    'feb': 2, 'febrero': 2, 'february': 2,
    'mar': 3, 'marzo': 3, 'march': 3,
    'abr': 4, 'abril': 4, 'april': 4,
    'may': 5, 'mayo': 5, 'may': 5,
    'jun': 6, 'junio': 6, 'june': 6,
    'jul': 7, 'julio': 7, 'july': 7,
    'ago': 8, 'agosto': 8, 'august': 8, 'aug': 8,
    'sep': 9, 'septiembre': 9, 'september': 9,
    'oct': 10, 'octubre': 10, 'october': 10,
    'nov': 11, 'noviembre': 11, 'november': 11,
    'dic': 12, 'diciembre': 12, 'december': 12, 'dec': 12
}

MONTH_DISPLAY = {
    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
}

def _get_cell_row(cell_ref: str) -> int:
    match = re.match(r"([A-Za-z]+)(\d+)", cell_ref)
    if match:
        return int(match.group(2))
    return 9999

def _compute_month_slots() -> list[dict]:
    """
    Compute the 7 month slots for the Teaser, desfased 2 months from today.
    Returns list of dicts with 'month_idx', 'year', 'label' sorted oldest-first.
    E.g. if today is May 2026 → Mar-26 backwards: Sep-25, Oct-25, Nov-25, Dec-25, Ene-26, Feb-26, Mar-26
    """
    from datetime import date
    from dateutil.relativedelta import relativedelta

    today = date.today()
    start = today - relativedelta(months=2)  # 2 months back from today
    slots = []
    for i in range(7):
        target = start - relativedelta(months=i)
        yr_suffix = str(target.year)[-2:]
        label = f"{MONTH_DISPLAY[target.month]} -{yr_suffix}"
        slots.append({
            "month_idx": target.month,
            "year": target.year,
            "label": label,
        })
    # Reverse so oldest is first (row 4 = oldest, row 10 = most recent)
    slots.reverse()
    return slots


def fill_template(template_path: str, output_path: str, data_list: list[dict], mapping: dict | None = None) -> str:
    """
    Fill an Excel template with extracted data using dynamic block assignment.
    Month rows are determined by the current date (desfased 2 months).
    Each data entry is placed in the row matching its extracted month/year.
    
    Args:
        template_path: Path to the .xlsx template
        output_path: Path where the filled .xlsx will be saved
        data_list: List of dicts of extracted financial data (multiple documents)
        mapping: Dict with 'blocks' array defining available cell positions
    
    Returns:
        Path to the generated output file
    """
    # Copy template to output
    shutil.copy2(template_path, output_path)
    
    wb = load_workbook(output_path)
    
    if mapping and "blocks" in mapping:
        # ── Dynamic block assignment mode ──
        
        # 1. Select the correct sheet
        sheet_name = mapping.get("sheet_name", "Bancos")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif sheet_name.upper() in wb.sheetnames:
            ws = wb[sheet_name.upper()]
        else:
            ws = wb.active
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet: {ws.title}")
        
        # 2. Compute the 7 dynamic month slots
        month_slots = _compute_month_slots()
        logger.info(f"Dynamic month slots: {[s['label'] for s in month_slots]}")
        
        # 3. Group data by account_name
        accounts = defaultdict(list)
        for data in data_list:
            acct = data.get("account_name", "").strip()
            if acct:
                accounts[acct].append(data)
        
        # 4. Sort accounts alphabetically for consistent ordering
        sorted_accounts = sorted(accounts.keys())
        
        # 5. Get available blocks from mapping
        blocks = mapping["blocks"]
        
        if len(sorted_accounts) > len(blocks):
            logger.warning(
                f"More unique accounts ({len(sorted_accounts)}) than available blocks ({len(blocks)}). "
                f"Extra accounts will be skipped."
            )
        
        # 6. Assign each unique account to the next available block
        for block_idx, account_name in enumerate(sorted_accounts):
            if block_idx >= len(blocks):
                logger.warning(f"No more blocks available for account '{account_name}', skipping.")
                break
            
            block = blocks[block_idx]
            base_cell = block.get("base")
            
            # Write account name to header cell
            if base_cell:
                ws[base_cell] = account_name
                logger.info(f"Assigned block {block_idx + 1} ({base_cell}) → {account_name}")
            
            # Get sorted lists of cells by their row coordinates (row 4 = slot 0, row 10 = slot 6)
            sorted_dep_cells = sorted(list(block.get("depositos", {}).values()), key=_get_cell_row)
            sorted_bal_cells = sorted(list(block.get("saldo_promedio", {}).values()), key=_get_cell_row)
            
            # Write month labels for ALL 7 slots (even empty ones)
            for slot_idx, slot in enumerate(month_slots):
                if slot_idx < len(sorted_dep_cells):
                    cell_obj = ws[sorted_dep_cells[slot_idx]]
                    label_col = cell_obj.column - 1
                    if label_col > 0:
                        ws.cell(row=cell_obj.row, column=label_col, value=slot["label"])
            
            # Place each data entry in the row that matches its month/year
            for data in accounts[account_name]:
                month_str = str(data.get("month", "")).lower().strip()[:3]
                data_month_idx = SPANISH_MONTHS.get(month_str, 0)
                try:
                    data_year = int(data.get("year", 0))
                except (ValueError, TypeError):
                    data_year = 0
                
                deposits = data.get("deposits")
                avg_balance = data.get("average_balance")
                
                # Find which slot this data belongs to
                target_slot = None
                for slot_idx, slot in enumerate(month_slots):
                    if slot["month_idx"] == data_month_idx and slot["year"] == data_year:
                        target_slot = slot_idx
                        break
                
                if target_slot is None:
                    logger.warning(
                        f"Data for {account_name} month={month_str} year={data_year} "
                        f"doesn't match any of the 7 expected slots, skipping."
                    )
                    continue
                
                if target_slot >= len(sorted_dep_cells):
                    continue
                
                # Write deposits
                if deposits is not None:
                    ws[sorted_dep_cells[target_slot]] = deposits
                
                # Write average balance
                if avg_balance is not None and target_slot < len(sorted_bal_cells):
                    ws[sorted_bal_cells[target_slot]] = avg_balance
    
    # Legacy support: old "Bancos" dict format (backwards compatible)
    elif mapping and "Bancos" in mapping:
        sheet_name = mapping.get("sheet_name", "Bancos")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        accounts = defaultdict(list)
        for data in data_list:
            acct = data.get("account_name", "").strip()
            if acct:
                accounts[acct].append(data)
        
        sorted_accounts = sorted(accounts.keys())
        blocks_list = list(mapping["Bancos"].values())
        
        for block_idx, account_name in enumerate(sorted_accounts):
            if block_idx >= len(blocks_list):
                break
            block = blocks_list[block_idx]
            base_cell = block.get("base")
            if base_cell:
                ws[base_cell] = account_name
            for data in accounts[account_name]:
                month = data.get("month", "").lower().strip()
                deposits = data.get("deposits")
                avg_balance = data.get("average_balance")
                if not month:
                    continue
                if deposits is not None:
                    dep_cell = block.get("depositos", {}).get(month)
                    if dep_cell:
                        ws[dep_cell] = deposits
                if avg_balance is not None:
                    bal_cell = block.get("saldo_promedio", {}).get(month)
                    if bal_cell:
                        ws[bal_cell] = avg_balance
    
    # Fallback for simple key-value mappings
    elif mapping:
        ws = wb.active
        if data_list:
            data = data_list[0]
            for key, cell_ref in mapping.items():
                if key in data:
                    ws[cell_ref] = data[key]
    
    wb.save(output_path)
    return output_path


def read_template_fields(template_path: str) -> dict:
    """
    Read an Excel template and return info about its structure.
    """
    wb = load_workbook(template_path, data_only=True)
    sheets_info = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets_info[sheet_name] = {
            "rows": ws.max_row,
            "cols": ws.max_column,
        }
    
    wb.close()
    return sheets_info
