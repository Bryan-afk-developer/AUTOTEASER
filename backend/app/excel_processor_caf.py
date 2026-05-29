"""
Excel Template Processor module.
Reads Excel templates and fills them with extracted financial data.
"""
import logging
from pathlib import Path
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def read_template_fields(template_path: str | Path) -> dict:
    """
    Read an Excel template and identify all fields/headers that need to be filled.
    
    Returns:
        dict with 'sheets' containing field info per sheet
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    wb = load_workbook(str(template_path), data_only=True)
    sheets_info = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fields = []
        headers = []
        
        # Strategy 1: Look for labeled cells (e.g., "Nombre:", "Saldo:")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip()
                    # Check if it's a label (ends with : or is a header)
                    if value.endswith(":") or value.endswith("="):
                        fields.append({
                            "label": value.rstrip(":= "),
                            "label_cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "value_cell": f"{get_column_letter(cell.column + 1)}{cell.row}",
                            "row": cell.row,
                            "col": cell.column
                        })
        
        # Strategy 2: Look for table headers (first row or first row with data)
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), max_col=ws.max_column):
            row_values = [cell.value for cell in row if cell.value]
            if len(row_values) >= 3:  # Likely a header row
                headers.append({
                    "row": row[0].row,
                    "columns": [
                        {
                            "name": str(cell.value).strip(),
                            "col_letter": get_column_letter(cell.column),
                            "col_index": cell.column
                        }
                        for cell in row if cell.value
                    ]
                })
        
        # Strategy 3: Look for placeholder markers like {field_name} or <<field_name>>
        placeholders = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip()
                    if (value.startswith("{") and value.endswith("}")) or \
                       (value.startswith("<<") and value.endswith(">>")):
                        clean_name = value.strip("{}< >")
                        placeholders.append({
                            "name": clean_name,
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "row": cell.row,
                            "col": cell.column
                        })
        
        sheets_info[sheet_name] = {
            "labeled_fields": fields,
            "table_headers": headers,
            "placeholders": placeholders,
            "dimensions": f"{ws.max_column} cols x {ws.max_row} rows"
        }
    
    wb.close()
    
    return {
        "template_file": template_path.name,
        "sheets": sheets_info,
        "sheet_names": wb.sheetnames
    }


def fill_template(
    template_path: str | Path,
    output_path: str | Path,
    data: dict,
    mapping: dict | None = None
) -> str:
    """
    Fill an Excel template with extracted financial data.
    
    Args:
        template_path: Path to the Excel template
        output_path: Path for the output filled Excel
        data: Extracted financial data dictionary
        mapping: Optional mapping of {cell_reference: data_key_path}
                 e.g., {"B2": "data.saldo_inicial", "B3": "data.saldo_final"}
    
    Returns:
        Path to the generated file
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    wb = load_workbook(str(template_path))
    
    if mapping:
        # Use explicit mapping
        _fill_with_mapping(wb, data, mapping)
    else:
        # Auto-fill based on template structure
        _auto_fill(wb, data)
        
    # --- RAW DUMP ---
    raw_dump = data.get("raw_text_dump")
    if raw_dump:
        ws_ocr = wb.create_sheet("Analíticas OCR")
        ws_ocr.append(["Año", "Línea Original (Extraída por OCR)"])
        for year, lines in raw_dump.items():
            for line in lines:
                ws_ocr.append([year, line])
        # Auto-adjust column width
        ws_ocr.column_dimensions['A'].width = 15
        ws_ocr.column_dimensions['B'].width = 120
    
    wb.save(str(output_path))
    wb.close()
    
    logger.info(f"Template filled and saved to: {output_path}")
    return str(output_path)


def _fill_with_mapping(wb: Workbook, data: dict, mapping: dict):
    """Fill cells based on explicit mapping. Supports flat mappings and nested CAF mappings."""
    
    # Check if it's the CAF nested mapping structure (either old lowercase or new literal style)
    is_caf_map = any(k.lower() in ["balance", "estado_resultados", "edo de resultados"] for k in mapping.keys())
    
    if is_caf_map:
        for categoria, anios in mapping.items():
            # Support both exact sheet names and previous mapped names
            if categoria == "balance":
                nombre_hoja = "Balance"
            elif categoria == "estado_resultados":
                nombre_hoja = "Edo de resultados"
            else:
                nombre_hoja = categoria
                
            if nombre_hoja not in wb.sheetnames:
                logger.warning(f"Sheet {nombre_hoja} not found in template, skipping {categoria}")
                continue
                
            ws = wb[nombre_hoja]
            
            # Automatically inject the lowest year into B3 based on the MAP so headers align perfectly
            years_in_map = [int(y) for y in anios.keys() if str(y).isdigit()]
            if years_in_map:
                lowest_year = min(years_in_map)
                ws["B3"] = lowest_year
                logger.debug(f"CAF Map: Auto-set Base Year {nombre_hoja}!B3 = {lowest_year}")
            
            # Get data key in a case-insensitive way
            categoria_data = {}
            for k, v in data.items():
                if k.lower() == categoria.lower():
                    categoria_data = v
                    break
                elif k.lower() == "estado_resultados" and categoria.lower() == "edo de resultados":
                    categoria_data = v
                    break
                elif k.lower() == "edo de resultados" and categoria.lower() == "estado_resultados":
                    categoria_data = v
                    break
            
            for anio, rubros in anios.items():
                # Support both string and integer keys for the year
                anio_data = {}
                if anio in categoria_data:
                    anio_data = categoria_data[anio]
                elif str(anio) in categoria_data:
                    anio_data = categoria_data[str(anio)]
                elif isinstance(categoria_data, dict):
                    for a_k, a_v in categoria_data.items():
                        if str(a_k) == str(anio):
                            anio_data = a_v
                            break
                            
                for rubro, coordenada_celda in rubros.items():
                    # Support case-insensitive rubro lookup
                    valor = None
                    if rubro in anio_data:
                        valor = anio_data[rubro]
                    elif isinstance(anio_data, dict):
                        for r_k, r_v in anio_data.items():
                            if r_k.lower() == rubro.lower():
                                valor = r_v
                                break
                    
                    try:
                        valor = float(valor) if valor is not None else 0.0
                    except (TypeError, ValueError):
                        valor = 0.0
                    
                    # Rubros que SIEMPRE deben ser negativos (son restas contables)
                    RUBROS_NEGATIVOS = {
                        # Balance
                        "depreciacion_acumulada",
                        # Estado de Resultados (todos son deducciones)
                        "costo_ventas",
                        "gastos_generales",
                        "gastos_administracion",
                        "gastos_financieros",
                        "productos_financieros",
                        "otros_gastos",
                        "otros_ingresos",
                        "impuestos",
                    }
                    if rubro in RUBROS_NEGATIVOS and valor > 0:
                        valor = -valor
                        
                    ws[coordenada_celda] = valor
                    logger.debug(f"CAF Map: Set {nombre_hoja}!{coordenada_celda} = {valor} (from {categoria}.{anio}.{rubro})")
        return

    # Fallback to standard flat mapping (Sheet1!B2: data.path)
    for cell_ref, data_path in mapping.items():
        if "!" in cell_ref:
            sheet_name, cell = cell_ref.split("!", 1)
            ws = wb[sheet_name]
        else:
            ws = wb.active
            cell = cell_ref
        
        value = _get_nested_value(data, data_path)
        if value is not None:
            ws[cell] = value
            logger.debug(f"Set {cell} = {value}")


def _auto_fill(wb: Workbook, data: dict):
    """
    Automatically fill template by matching field names with data keys.
    Uses fuzzy matching to handle variations in naming.
    """
    flat_data = _flatten_dict(data)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip()
                    
                    # Handle placeholders: {field_name} or <<field_name>>
                    if (value.startswith("{") and value.endswith("}")) or \
                       (value.startswith("<<") and value.endswith(">>")):
                        field_name = value.strip("{}< >").lower()
                        matched_value = _find_matching_value(field_name, flat_data)
                        if matched_value is not None:
                            cell.value = matched_value
                    
                    # Handle labels (with or without colons)
                    else:
                        label = value.rstrip(":=").strip().lower()
                        # Ignore very short strings or obvious non-labels
                        if len(label) > 2 and not label.isnumeric():
                            # Remove leading numbers/bullets like "1 Caja" -> "Caja"
                            import re
                            clean_label = re.sub(r'^\d+[\s\.\-\)]+', '', label).strip()
                            
                            matched_value = _find_matching_value(clean_label, flat_data)
                            if matched_value is not None:
                                # Search to the right for the first available empty numeric column
                                # We scan up to 15 columns to the right
                                for offset in range(1, 16):
                                    target_col = cell.column + offset
                                    current_val = ws.cell(row=cell.row, column=target_col).value
                                    
                                    # If the cell is empty, 0, or None, we can write here!
                                    # We also want to skip columns that have strings like "%", "$", or account codes
                                    if current_val is None or current_val == 0 or current_val == "":
                                        ws.cell(row=cell.row, column=target_col, value=matched_value)
                                        break
                                    elif isinstance(current_val, str) and (current_val.strip() == "$" or current_val.strip() == "%"):
                                        # Skip formatting columns
                                        continue
                                    elif isinstance(current_val, str) and len(current_val) > 4:
                                        # Skip account codes like "001701"
                                        continue
                                    else:
                                        # Cell is occupied with actual data from a previous year, keep looking right
                                        continue


def fill_template_with_movements(
    template_path: str | Path,
    output_path: str | Path,
    header_data: dict,
    movements: list[dict],
    sheet_name: str | None = None,
    data_start_row: int | None = None
) -> str:
    """
    Fill a template that has both header fields and a movements table.
    
    Args:
        template_path: Path to template
        output_path: Path for output
        header_data: Dict with header-level data (name, dates, totals, etc.)
        movements: List of movement dicts
        sheet_name: Target sheet name
        data_start_row: Row where movement data should start
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    
    wb = load_workbook(str(template_path))
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # Fill header data
    flat_header = _flatten_dict(header_data)
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), max_col=ws.max_column):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                value = cell.value.strip()
                if (value.startswith("{") and value.endswith("}")) or \
                   (value.startswith("<<") and value.endswith(">>")):
                    field_name = value.strip("{}< >").lower()
                    matched = _find_matching_value(field_name, flat_header)
                    if matched is not None:
                        cell.value = matched
    
    # Find the table header row to determine column mapping
    if not data_start_row:
        data_start_row = _find_table_start(ws)
    
    if data_start_row and movements:
        # Get column headers
        header_row = data_start_row - 1
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col).value
            if header_val:
                col_mapping[str(header_val).strip().lower()] = col
        
        # Write movements
        for i, mov in enumerate(movements):
            row_num = data_start_row + i
            
            # Insert row if needed (preserve formatting)
            if i > 0 and row_num <= ws.max_row:
                ws.insert_rows(row_num)
            
            flat_mov = _flatten_dict(mov)
            for header_name, col_idx in col_mapping.items():
                value = _find_matching_value(header_name, flat_mov)
                if value is not None:
                    ws.cell(row=row_num, column=col_idx, value=value)
    
    wb.save(str(output_path))
    wb.close()
    
    return str(output_path)


def create_default_template(output_path: str | Path, doc_type: str = "estado_de_cuenta") -> str:
    """
    Create a default Excel template based on document type.
    """
    wb = Workbook()
    ws = wb.active
    
    if doc_type == "estado_de_cuenta":
        ws.title = "Estado de Cuenta"
        
        # Header section
        headers = [
            ("A1", "Institución:"), ("B1", "{institucion}"),
            ("A2", "Titular:"), ("B2", "{titular}"),
            ("A3", "No. Cuenta:"), ("B3", "{numero_cuenta}"),
            ("A4", "Periodo:"), ("B4", "{periodo}"),
            ("A5", "Fecha Corte:"), ("B5", "{fecha_corte}"),
            ("A6", "Moneda:"), ("B6", "{moneda}"),
            ("A8", "Saldo Inicial:"), ("B8", "{saldo_inicial}"),
            ("A9", "Saldo Final:"), ("B9", "{saldo_final}"),
            ("A10", "Total Depósitos:"), ("B10", "{total_depositos}"),
            ("A11", "Total Retiros:"), ("B11", "{total_retiros}"),
            ("A12", "Comisiones:"), ("B12", "{total_comisiones}"),
            ("A13", "Intereses:"), ("B13", "{intereses_generados}"),
        ]
        
        for cell, value in headers:
            ws[cell] = value
        
        # Movements table header
        mov_headers = ["Fecha", "Concepto", "Referencia", "Cargo", "Abono", "Saldo"]
        for i, h in enumerate(mov_headers, 1):
            ws.cell(row=15, column=i, value=h)
        
        # Set column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
    elif doc_type == "balance_general":
        ws.title = "Balance General"
        
        rows = [
            ("A1", "BALANCE GENERAL"),
            ("A2", "Empresa:"), ("B2", "{empresa}"),
            ("A3", "Periodo:"), ("B3", "{periodo}"),
            ("A4", "Fecha:"), ("B4", "{fecha}"),
            ("A6", "ACTIVO"),
            ("A7", "Activo Circulante"),
            ("A8", "  Caja y Bancos:"), ("B8", "{caja_y_bancos}"),
            ("A9", "  Clientes:"), ("B9", "{clientes}"),
            ("A10", "  Inventarios:"), ("B10", "{inventarios}"),
            ("A11", "  Cuentas por Cobrar:"), ("B11", "{cuentas_por_cobrar}"),
            ("A12", "Total Activo Circulante:"), ("B12", "{total_activo_circulante}"),
            ("A14", "Activo Fijo"),
            ("A15", "  Terrenos:"), ("B15", "{terrenos}"),
            ("A16", "  Edificios:"), ("B16", "{edificios}"),
            ("A17", "  Maquinaria y Equipo:"), ("B17", "{maquinaria_y_equipo}"),
            ("A18", "  Dep. Acumulada:"), ("B18", "{depreciacion_acumulada}"),
            ("A19", "Total Activo Fijo:"), ("B19", "{total_activo_fijo}"),
            ("A21", "TOTAL ACTIVO:"), ("B21", "{total_activo}"),
            ("A23", "PASIVO"),
            ("A24", "Pasivo Circulante"),
            ("A25", "  Proveedores:"), ("B25", "{proveedores}"),
            ("A26", "  Cuentas por Pagar:"), ("B26", "{cuentas_por_pagar}"),
            ("A27", "  Impuestos por Pagar:"), ("B27", "{impuestos_por_pagar}"),
            ("A28", "Total Pasivo Circulante:"), ("B28", "{total_pasivo_circulante}"),
            ("A30", "TOTAL PASIVO:"), ("B30", "{total_pasivo}"),
            ("A32", "CAPITAL CONTABLE"),
            ("A33", "  Capital Social:"), ("B33", "{capital_social}"),
            ("A34", "  Resultado del Ejercicio:"), ("B34", "{resultado_del_ejercicio}"),
            ("A35", "Total Capital Contable:"), ("B35", "{total_capital_contable}"),
            ("A37", "PASIVO + CAPITAL:"), ("B37", "{total_pasivo_mas_capital}"),
        ]
        
        for cell, value in rows:
            ws[cell] = value
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    output_path = Path(output_path)
    wb.save(str(output_path))
    wb.close()
    
    return str(output_path)


# ── Helper functions ──────────────────────────────────────────────────────

def _flatten_dict(d: dict, parent_key: str = "", sep: str = ".") -> dict:
    """Flatten a nested dictionary."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key.lower(), v))
    return dict(items)


def _get_nested_value(data: dict, path: str):
    """Get a value from a nested dict using dot notation path."""
    keys = path.split(".")
    current = data
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            return None
    return current


def _find_matching_value(field_name: str, flat_data: dict):
    """Find the best matching value for a field name in flattened data."""
    field_lower = field_name.lower().replace(" ", "_").replace("-", "_")
    
    # Exact match
    if field_lower in flat_data:
        return flat_data[field_lower]
    
    # Partial match - field is at the end of a key path
    for key, value in flat_data.items():
        if key.endswith(f".{field_lower}") or key == field_lower:
            return value
    
    # Fuzzy partial match
    for key, value in flat_data.items():
        key_parts = key.split(".")
        if field_lower in key_parts:
            return value
    
    # Contains match
    for key, value in flat_data.items():
        if field_lower in key or key in field_lower:
            return value
    
    return None


def _find_table_start(ws) -> int | None:
    """Find the row where a data table starts (row after headers)."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        row_values = [cell.value for cell in row if cell.value is not None]
        if len(row_values) >= 3:
            # Check if this looks like a header row
            all_strings = all(isinstance(v, str) for v in row_values)
            if all_strings:
                return row[0].row + 1
    return None

# ══════════════════════════════════════════════════════════════════
# ANALYTICS SHEET INJECTION (v2 - Verified Groups)
# ══════════════════════════════════════════════════════════════════

def fill_analytics_sheet(
    workbook_path: str | Path,
    output_path: str | Path,
    analytics_data: dict,
) -> str:
    """
    Create/overwrite an 'Analíticas' sheet in the CAF workbook with
    verified parent→children financial analytics.

    Format per group:
        CLIENTES                    17,453,304
        -PORKCHARRON S. DE R.L...      690,823
        -CHATA SNACK'S S DE RL...      439,950
        ...
        =VERIFICACIÓN               17,453,304  ✓

    Args:
        workbook_path: Path to the CAF Excel template/workbook
        output_path: Path for the output filled Excel
        analytics_data: Result from analytics_parser.parse_analytics() v2

    Returns:
        Path to the generated file
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    workbook_path = Path(workbook_path)
    output_path = Path(output_path)

    wb = load_workbook(str(workbook_path))

    # Normalize input: we want a dict { year: analytics_data }
    if not isinstance(analytics_data, dict):
        analytics_dict = {"2024": analytics_data} if analytics_data else {}
    elif "success" in analytics_data and "groups" in analytics_data:
        year = analytics_data.get("year", "2024")
        analytics_dict = {year: analytics_data}
    else:
        analytics_dict = analytics_data

    # Clean up old default sheets
    for name_candidate in ["Analíticas", "Analiticas"]:
        if name_candidate in wb.sheetnames:
            del wb[name_candidate]

    for year, data in analytics_dict.items():
        if not data or not data.get("groups"):
            continue
            
        sheet_name = f"Analíticas {year}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        ws = wb.create_sheet(title=sheet_name)
        groups = data.get("groups", [])
        
        # ── Styles ──
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

        parent_font = Font(name='Calibri', size=10, bold=True, color='1F3864')
        parent_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        child_font = Font(name='Calibri', size=10, color='333333')

        verify_ok_font = Font(name='Calibri', size=10, bold=True, color='006100')
        verify_ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        verify_fail_font = Font(name='Calibri', size=10, bold=True, color='9C0006')
        verify_fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        amount_align = Alignment(horizontal='right')
        num_fmt = '#,##0.00'

        # ── Column widths ──
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12

        # ── Header row ──
        row = 1
        for col, val in [(1, "Concepto"), (2, f"Monto ({year})"), (3, "Estado")]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center') if col == 3 else (amount_align if col == 2 else Alignment(horizontal='left'))

        row = 3  # Skip a blank row

        # ── Write each group ──
        verified_total = 0
        failed_total = 0

        for group in groups:
            parent_concept = group['parent_concept']
            parent_total = group['parent_total']
            children = group['children']
            verified = group['verified']
            children_sum = group['children_sum']

            # Parent row
            cell_a = ws.cell(row=row, column=1, value=parent_concept)
            cell_a.font = parent_font
            cell_a.fill = parent_fill

            cell_b = ws.cell(row=row, column=2, value=parent_total)
            cell_b.font = parent_font
            cell_b.fill = parent_fill
            cell_b.alignment = amount_align
            cell_b.number_format = num_fmt
            row += 1

            # Children rows
            first_child_row = row
            for child in children:
                child_label = f"  - {child['concept']}"

                cell_a = ws.cell(row=row, column=1, value=child_label)
                cell_a.font = child_font

                cell_b = ws.cell(row=row, column=2, value=child['amount'])
                cell_b.font = child_font
                cell_b.alignment = amount_align
                cell_b.number_format = num_fmt
                row += 1
            last_child_row = row - 1

            # Verification row
            if children:
                if verified:
                    label = f"  = SUMA VERIFICADA"
                    v_font = verify_ok_font
                    v_fill = verify_ok_fill
                    status = "OK"
                    verified_total += 1
                else:
                    diff = group.get('diff', 0)
                    label = f"  = SUMA (dif: {diff:,.2f})"
                    v_font = verify_fail_font
                    v_fill = verify_fail_fill
                    status = "ERROR"
                    failed_total += 1

                cell_a = ws.cell(row=row, column=1, value=label)
                cell_a.font = v_font
                cell_a.fill = v_fill

                # Use Excel Formula for the sum
                formula = f"=SUM(B{first_child_row}:B{last_child_row})"
                cell_b = ws.cell(row=row, column=2, value=formula)
                cell_b.font = v_font
                cell_b.fill = v_fill
                cell_b.alignment = amount_align
                cell_b.number_format = num_fmt

                cell_c = ws.cell(row=row, column=3, value=status)
                cell_c.font = v_font
                cell_c.fill = v_fill
                cell_c.alignment = Alignment(horizontal='center')
                row += 1

            row += 1  # Blank row between groups

        # ── Summary footer ──
        row += 1
        for col, val in [(1, "Resumen de Verificacion"), (2, "Grupos"), (3, "Estado")]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center') if col > 1 else Alignment(horizontal='left')
        row += 1

        ws.cell(row=row, column=1, value="Verificados correctamente").font = verify_ok_font
        ws.cell(row=row, column=2, value=verified_total).font = verify_ok_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value="OK").font = verify_ok_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row=row, column=1, value="Con diferencia").font = verify_fail_font
        ws.cell(row=row, column=2, value=failed_total).font = verify_fail_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value="ERROR" if failed_total > 0 else "-").font = verify_fail_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

    wb.save(str(output_path))
    wb.close()
    return str(output_path)


    # ── Styles ──
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')

    parent_font = Font(name='Calibri', size=10, bold=True, color='1F3864')
    parent_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    child_font = Font(name='Calibri', size=10, color='333333')

    verify_ok_font = Font(name='Calibri', size=10, bold=True, color='006100')
    verify_ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    verify_fail_font = Font(name='Calibri', size=10, bold=True, color='9C0006')
    verify_fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    amount_align = Alignment(horizontal='right')
    num_fmt = '#,##0.00'
    num_fmt_int = '#,##0'

    # ── Column widths ──
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12

    # ── Header row ──
    row = 1
    for col, val in [(1, "Concepto"), (2, f"Monto ({year})"), (3, "Estado")]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center') if col == 3 else (amount_align if col == 2 else Alignment(horizontal='left'))

    row = 3  # Skip a blank row

    # ── Write each group ──
    verified_total = 0
    failed_total = 0

    for group in groups:
        parent_concept = group['parent_concept']
        parent_total = group['parent_total']
        children = group['children']
        verified = group['verified']
        children_sum = group['children_sum']

        # Parent row
        cell_a = ws.cell(row=row, column=1, value=parent_concept)
        cell_a.font = parent_font
        cell_a.fill = parent_fill

        cell_b = ws.cell(row=row, column=2, value=parent_total)
        cell_b.font = parent_font
        cell_b.fill = parent_fill
        cell_b.alignment = amount_align
        cell_b.number_format = num_fmt
        row += 1

        # Children rows
        for child in children:
            child_label = f"  - {child['concept']}"

            cell_a = ws.cell(row=row, column=1, value=child_label)
            cell_a.font = child_font

            cell_b = ws.cell(row=row, column=2, value=child['amount'])
            cell_b.font = child_font
            cell_b.alignment = amount_align
            cell_b.number_format = num_fmt
            row += 1

        # Verification row
        if children:
            if verified:
                label = f"  = SUMA VERIFICADA"
                v_font = verify_ok_font
                v_fill = verify_ok_fill
                status = "OK"
                verified_total += 1
            else:
                diff = group.get('diff', 0)
                label = f"  = SUMA (dif: {diff:,.2f})"
                v_font = verify_fail_font
                v_fill = verify_fail_fill
                status = "ERROR"
                failed_total += 1

            cell_a = ws.cell(row=row, column=1, value=label)
            cell_a.font = v_font
            cell_a.fill = v_fill

            cell_b = ws.cell(row=row, column=2, value=children_sum)
            cell_b.font = v_font
            cell_b.fill = v_fill
            cell_b.alignment = amount_align
            cell_b.number_format = num_fmt

            cell_c = ws.cell(row=row, column=3, value=status)
            cell_c.font = v_font
            cell_c.fill = v_fill
            cell_c.alignment = Alignment(horizontal='center')
            row += 1

        row += 1  # Blank row between groups

    # ── Summary footer ──
    row += 1
    for col, val in [(1, "Resumen de Verificacion"), (2, "Grupos"), (3, "Estado")]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center') if col > 1 else Alignment(horizontal='left')
    row += 1

    ws.cell(row=row, column=1, value="Verificados correctamente").font = verify_ok_font
    ws.cell(row=row, column=2, value=verified_total).font = verify_ok_font
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3, value="OK").font = verify_ok_font
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    row += 1

    ws.cell(row=row, column=1, value="Con diferencia").font = verify_fail_font
    ws.cell(row=row, column=2, value=failed_total).font = verify_fail_font
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3, value="ERROR" if failed_total > 0 else "-").font = verify_fail_font
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')

    wb.save(str(output_path))
    wb.close()

    logger.info(f"Analytics sheet written: {len(groups)} groups, "
                f"{verified_total} verified, {failed_total} failed -> {output_path}")
    return str(output_path)


